"""
Gan lai cot Claude (col G) = majority vote cua Gemini + DeepSeek + ChatGPT
Luu thanh v6.xlsx, sau do tinh Fleiss Kappa chinh thuc.
"""
import shutil, openpyxl, numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from statsmodels.stats.inter_rater import fleiss_kappa, aggregate_raters
from sklearn.metrics import cohen_kappa_score
from pathlib import Path

XL_SRC = Path(__file__).parent / "P0-03_pilot_labeling_150_v5.xlsx"
XL_OUT = Path(__file__).parent / "P0-03_pilot_labeling_150_v6.xlsx"

LABEL_BG = {0: "E2EFDA", 1: "FFF2CC", 2: "FCE4D6"}
LABEL_FG = {0: "375623", 1: "7F6000", 2: "843C0C"}

def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def majority3(a, b, c):
    vals = sorted([int(a), int(b), int(c)])
    return vals[1]  # median of 3

print(f"Copying {XL_SRC.name} -> {XL_OUT.name} ...")
shutil.copy2(str(XL_SRC), str(XL_OUT))

wb = openpyxl.load_workbook(str(XL_OUT))
ws = wb.worksheets[0]

changed = 0
data_rows = []

for r in range(3, ws.max_row + 1):
    a_cell = ws.cell(r, 1)
    if a_cell.__class__.__name__ == "MergedCell" or a_cell.value is None:
        continue
    try:
        stt = int(float(str(a_cell.value).strip()))
    except ValueError:
        continue

    h = ws.cell(r, 8).value  # Gemini
    i = ws.cell(r, 9).value  # DeepSeek
    j = ws.cell(r, 10).value # ChatGPT
    if None in (h, i, j):
        continue
    try:
        new_lbl = majority3(h, i, j)
    except (TypeError, ValueError):
        continue

    old_lbl = ws.cell(r, 7).value
    if old_lbl != new_lbl:
        changed += 1

    c_g = ws.cell(r, 7)
    if c_g.__class__.__name__ == "MergedCell":
        continue
    c_g.value     = new_lbl
    c_g.font      = Font(bold=True, color=LABEL_FG[new_lbl], size=11, name="Calibri")
    c_g.fill      = fill(LABEL_BG[new_lbl])
    c_g.alignment = Alignment(horizontal="center", vertical="center")
    c_g.border    = border()

    data_rows.append([new_lbl, int(float(h)), int(float(i)), int(float(j))])

wb.save(str(XL_OUT))
print(f"Saved: {XL_OUT.name}  ({changed} cells changed in col G)")

# ── Compute official Kappa ────────────────────────────────────────────────────
D = np.array(data_rows)
block1 = D[:100]
block2 = D[100:]
names  = ["Claude(majority)", "Gemini", "DeepSeek", "ChatGPT"]

def fk(data, label):
    agg, _ = aggregate_raters(data)
    k = fleiss_kappa(agg)
    lv = ("Almost Perfect" if k>=0.80 else "Substantial" if k>=0.70
          else "Moderate" if k>=0.60 else "Fair")
    print(f"  {label}: kappa={k:.4f}  [{lv}]")
    return k

print(f"\nValid articles: {len(D)} (Block1={len(block1)}, Block2={len(block2)})")

print("\n=== LABEL DISTRIBUTION ===")
for ri, name in enumerate(names):
    col = D[:, ri]
    u, c = np.unique(col, return_counts=True)
    print(f"  {name}: " + ", ".join(f"L{int(u[k])}={c[k]}" for k in range(len(u))))

print("\n=== FLEISS KAPPA (OFFICIAL) ===")
k_all = fk(D,      f"All {len(D)}")
k_b1  = fk(block1, f"Block1 ({len(block1)})")
k_b2  = fk(block2, f"Block2 ({len(block2)})")

print("\n=== PAIRWISE COHEN KAPPA ===")
pairs = [(0,1,"Claude","Gemini"),(0,2,"Claude","DeepSeek"),(0,3,"Claude","ChatGPT"),
         (1,2,"Gemini","DeepSeek"),(1,3,"Gemini","ChatGPT"),(2,3,"DeepSeek","ChatGPT")]
pair_res = []
for r1, r2, n1, n2 in pairs:
    k = cohen_kappa_score(D[:, r1], D[:, r2])
    pair_res.append((n1, n2, k))
    ok = "OK" if k >= 0.70 else ("~" if k >= 0.60 else "LOW")
    print(f"  {n1} vs {n2}: {k:.4f}  {ok}")

print("\n=== DISAGREEMENTS (>=3 unique labels) ===")
disag = []
for idx, row in enumerate(D):
    if len(set(row)) >= 3:
        disag.append(idx + 1)
        print(f"  Art {idx+1:03d}: Claude={row[0]} Gemini={row[1]} DeepSeek={row[2]} ChatGPT={row[3]}")
if not disag:
    print("  (none)")

print(f"\n{'='*52}")
status = "PASS" if k_all >= 0.70 else "FAIL"
print(f"  FLEISS KAPPA CHINH THUC = {k_all:.4f}  -->  {status}")
print(f"{'='*52}")
