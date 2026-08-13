"""
P0-03 -- Gan lai cot Claude (col G) trong v4.xlsx theo classifier v3.

Thay doi so voi v2:
  1. Doc title+content truc tiep tu Excel (khong rebuild tu JSON) -> dung bai dung
  2. Ha nguong Label 1: m_score >= 1 la du (bo dieu kien aerospace)
  3. FUTURE_HEDGE chi ap cho HIGH -> MEDIUM downgrade (khong ngan Label 1)
  4. QT-03 (non-aerospace giam 1 bac) van giu
"""

import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path

XL_IN  = Path(__file__).parent / "P0-03_pilot_labeling_150_v4.xlsx"
XL_OUT = Path(__file__).parent / "P0-03_pilot_labeling_150_v4.xlsx"

# ── Pattern lists (same as v2) ────────────────────────────────────────────────

STRIKE_ACTIVE = [
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day\s*\d|launched|action\s+begins)\b',
    r'\b(workers?|dockworkers?|longshoremen|truckers?|employees?)\s+(are\s+)?(on\s+strike|walk\w+\s+out|went\s+on\s+strike|began\s+strik\w+)\b',
    r'\bwalkout\s+(begins|started|underway)\b',
    r'\bwork\s+stoppage\s+(begins|started|underway|in\s+effect)\b',
    r'\bpicket\s+line\b',
    r'\bstrikers?\s+(block|prevent|halt|shut)\b',
    r'\bstrike\s+(halt\w*|shut\w*|crippl\w*|disrupt\w*)\b',
]
RED_SEA_ACTIVE = [
    r'\b(houthi\w*)\s+(attack\w*|fire\w*|hit\s+|struck|seized|hijack\w*)\b',
    r'\b(attack\w*|fire\w*|struck|seized|hijack\w*)\b.{0,60}\b(red\s+sea|gulf\s+of\s+aden)\b',
    r'\b(maersk|hapag|msc|cma\s+cgm|evergreen|cosco|zim|yang\s+ming)\s+\w*\s*(suspend\w*|halt\w*|divert\w*|reroute\w*|avoid\w*|skip\w*)\b',
    r'\b(carrier\w*|shipping\s+line\w*|ship\w*)\s+\w*\s*(suspend\w*|halt\w*|divert\w*|reroute\w*)\b.{0,80}\b(red\s+sea|suez)\b',
    r'\bships?\s+(divert\w*|reroute\w*)\b.{0,60}\b(cape|good\s+hope|africa)\b',
    r'\bvoyage\s+around\s+(africa|cape)\b',
]
PORT_CLOSED = [
    r'\b(port|terminal|harbor)\s+(is\s+)?(clos\w+|shut\s*down|halt\w*|suspend\w*|block\w*)\b',
    r'\b(clos\w+|shut\s*down|halt\w*|suspend\w*)\s+(port|terminal|operations?)\b',
    r'\bport\s+of\s+\w+\s+(clos\w+|shut\w*|halt\w*)\b',
]
PLANT_CLOSED = [
    r'\b(plant|factory|facilit\w+|warehouse|assembly)\s+(is\s+)?(shut\w*|clos\w*|halt\w*|idl\w+|suspend\w*)\b',
    r'\b(shut\w*|clos\w*|halt\w*|suspend\w*)\s+(plant|factory|facilit\w+|production|operations?)\b',
    r'\bproduction\s+(halt\w*|suspend\w*|shut\w*|stop\w*)\b',
]
BANKRUPTCY_ACTIVE = [
    r'\bfil\w+\s+for\s+(bankruptcy|chapter\s+11|insolvenc\w*)\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\bceas\w+\s+(operat\w*|trading|business)\b',
    r'\bwent\s+(bankrupt|into\s+administration|insolvent)\b',
    r'\bdefault\w*\s+on\s+(debt\w*|payment\w*|loan\w*|bond\w*)\b',
]
DISASTER_ACTIVE = [
    r'\b(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami|cyclone|tornado)\s+\w{0,15}\s*(hit\w*|struck|devastat\w*|destroy\w*|slam\w*|batter\w*|sweep\w*|ravag\w*)\b',
    r'\b(hit|struck|devastat\w*|destroy\w*|slam\w*)\s+by\s+(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami)\b',
    r'\b(flood\w*|earthquake|fire)\s+(halt\w*|shut\w*|clos\w*|disrupt\w*|damage\w*)\s+(port|factory|plant|facilit\w+|road|bridge|infra)\b',
]
SANCTION_ACTIVE = [
    r'\b(sanction\w*|embargo|ban|restriction\w*)\s+(now\s+)?(in\s+effect|effective|enforced|imposed|applied|implement\w*)\b',
    r'\b(us|eu|un|uk)\s+(imposes?|imposed|enacted)\s+(sanction\w*|ban|embargo)\b',
    r'\bnew\s+sanction\w*\s+on\b',
]
LOCKDOWN_ACTIVE = [
    r'\b(lockdown|lock\s+down)\s+(in\s+effect|imposed|announced|begins|started|underway)\b',
    r'\b(city|region|province|district|area)\s+(lock\w+|shut\w+|seal\w+)\b',
    r'\bcovid\s*-?\s*19?\s+(lockdown|shutdown|closure)\b',
    r'\bzero\s*-?\s*covid\s+(policy|lockdown|measures?)\b',
]
SHORTAGE_CONFIRMED = [
    r'\b(stockout|out\s+of\s+stock|ran\s+out\s+of|depleted)\b',
    r'\b(critical|acute|severe)\s+shortage\b',
    r'\bshortage\s+(hit\w*|affect\w*|forc\w*|caus\w*)\b.{0,60}\b(production|assembly|manufactur)\b',
]
CARGO_SEIZED = [
    r'\b(cargo|vessel|ship|container)\s+(seized|detained|impounded|confiscated|blocked)\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*)\b',
    r'\bever\s*given\b',
]
SEVERE_DELAY = [
    r'\bdelays?\s+of\s+(several\s+)?(weeks?|months?)\b',
    r'\b(weeks?|months?)\s+of\s+delay\w*\b',
    r'\b\d+\s*-?\s*(week|month)\s+delay\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?|containers?)\s+(wait\w*|strand\w*|back\w*log\w*|anchor\w*|queue\w*)\b',
]

ALL_HIGH_PATTERNS = (
    STRIKE_ACTIVE + RED_SEA_ACTIVE + PORT_CLOSED + PLANT_CLOSED +
    BANKRUPTCY_ACTIVE + DISASTER_ACTIVE + SANCTION_ACTIVE +
    LOCKDOWN_ACTIVE + SHORTAGE_CONFIRMED + CARGO_SEIZED + SEVERE_DELAY
)

MEDIUM_PATTERNS = [
    r'\b(contract\s+talks?|labor\s+negotiat\w*|union\s+negotiat\w*|collective\s+bargaining)\b',
    r'\bstrike\s+(threat|vote|authoriz\w*|warning|loom\w*|possible|risk|could|may)\b',
    r'\b(workers?|union)\s+(vote|threaten|warn\w*|consider\w*|prepar\w*)\s+\w{0,20}\s*(strike|walkout|action)\b',
    r'\bstrike\s+authoriz\w+\b',
    r'\bcontract\s+(expir\w*|deadline|negotiat\w*|impasse|breakdown)\b',
    r'\b(warn\w*|alert\w*|caution\w*|advis\w*)\b.{0,60}\b(red\s+sea|gulf\s+of\s+aden|houthi)\b',
    r'\b(houthi\w*)\s+(threat\w*|warn\w*|target\w*|could\s+attack)\b',
    r'\binsurance\s+premium\w*\s+(rise\w*|increas\w*|surge\w*)\b.{0,60}\b(red\s+sea|gulf)\b',
    r'\b(propos\w+|plan\w*|consider\w*|draft\w*|mulling|weighing)\s+(tariff|sanction\w*|restriction|ban|levy)\b',
    r'\btariff\s+(hike|increase|propos\w*|threat\w*|possible|could)\b',
    r'\btrade\s+(war|dispute|tension\w*|conflict|friction)\b',
    r'\b(us|eu|un|uk)\s+(consider\w*|plan\w*|weigh\w*|discuss\w*)\s+(sanction\w*|ban|tariff)\b',
    r'\b(storm|typhoon|hurricane|flood|cyclone)\s+(approach\w*|threaten\w*|head\w+\s+toward|forecast\w*|warn\w*|watch|warning|expected\s+to\s+hit)\b',
    r'\bport\s+congestion\b',
    r'\b(increas\w+|grow\w+|ris\w+|worsening|escalat\w+)\s+(congestion|delay\w*|backlog)\b',
    r'\bshipping\s+(delay\w*|backlog|bottleneck|disruption)\b',
    r'\b(risk|threat|fear\w*|concern\w*)\s+(of\s+)?(shortage|shortfall|disruption|delay)\b',
    r'\bshortage\s+(risk|concern\w*|fear\w*|loom\w*|possible|potential|ahead)\b',
    r'\b(tight|thin|low|lean)\s+(supply|inventory|stock)\b',
    r'\binventory\s+(strain|pressure|concern|crunch)\b',
    r'\bsupplier\s+(issue\w*|problem\w*|concern\w*|strain\w*|stress\w*|challeng\w*|struggl\w*|distress\w*)\b',
    r'\b(financial\s+distress|cash\s+crunch|liquidity\s+crisis)\b',
    r'\b(longer|extended|stretched|growing)\s+lead\s+time\w*\b',
    r'\blead\s+time\w*\s+(increase\w*|grow\w*|worsen\w*|stretch\w*)\b',
    r'\b(geopolitical|trade)\s+(tension\w*|uncertainty|instability|dispute\w*)\b',
    r'\bsupply\s+chain\s+(risk\w*|vulnerab\w*|concern\w*|warn\w*)\b',
    r'\bpanama\s+canal\s+(restrict\w*|drought\w*|low\w*\s+water|limit\w*|backlog)\b',
    # Them: canh bao chinh sach, thue quan da ap
    r'\btariff\w*\s+(impact\w*|effect\w*|cost\w*|burden\w*)\b',
    r'\b(new|additional|higher)\s+tariff\w*\b',
    r'\bsupply\s+chain\s+(disruption\w*|challeng\w*|pressur\w*|strain\w*)\b',
    r'\bfreight\s+(rate\w*|cost\w*)\s+(rise\w*|surge\w*|spike\w*|soar\w*|increas\w*)\b',
    r'\bshipping\s+(cost\w*|rate\w*)\s+(rise\w*|surge\w*|spike\w*|soar\w*|increas\w*)\b',
    r'\b(labor|port)\s+(dispute\w*|unrest\w*|tension\w*)\b',
]

NEUTRAL_HARD = [
    r'\bartificial\s+intelligence\b', r'\bmachine\s+learning\b',
    r'\bdigital\s+(transform\w*|solution\w*|platform\w*)\b',
    r'\bblockchain\b', r'\binternet\s+of\s+things\b',
    r'\bwebinar\b', r'\bpodcast\b', r'\bwhitepaper\b',
    r'\b\d{4}\s+annual\s+report\b',
    r'\baward\w*\s+(winner|recipient|ceremony)\b',
    r'\bnew\s+(ceo|cfo|coo|vp|president|director|officer)\b',
    r'\bjoint\s+venture\s+(formed|announced|created|finalized)\b',
    r'\bmarket\s+(share|size|growth|forecast|outlook)\b',
    r'\brevenue\s+(grow\w*|increas\w*|reach\w*)\b',
    r'\bprofit\s+(rise\w*|increas\w*|beat\w*|outlook)\b',
    r'\bipo\b', r'\bfunding\s+round\b',
]

AEROSPACE_RELATED = [
    r'\baerospace\b', r'\baviation\b', r'\baircraft\b', r'\bairline\b',
    r'\bairport\b', r'\bairfreight\b', r'\bair\s+cargo\b', r'\bair\s+freight\b',
    r'\bboeing\b', r'\bairbus\b', r'\bdefense\b', r'\bmilitary\b',
    r'\bport\b', r'\bshipping\b', r'\bfreight\b', r'\bcontainer\b',
    r'\bsuez\b', r'\bpanama\b', r'\bred\s+sea\b', r'\bblack\s+sea\b',
    r'\btranspacific\b', r'\btransatlantic\b',
    r'\bsemiconductor\b', r'\bchip\s+shortage\b',
    r'\benergy\s+(supply|crisis|shortage)\b',
    r'\bsteel\b', r'\bcopper\b', r'\brare\s+earth\b',
]
NON_AEROSPACE_HARD = [
    r'\bautomotive\b', r'\bauto\s+(industry|maker|plant)\b',
    r'\belectric\s+vehicle\b', r'\bev\s+(battery|maker)\b',
    r'\bfarm\w*\b', r'\bagricultur\w*\b', r'\bgrain\b', r'\bwheat\b',
    r'\bsoybeans?\b', r'\bfertilizer\b',
    r'\bpharmaceutical\b', r'\bdrug\s+(supply|shortage)\b',
    r'\bfood\s+(supply|chain|shortage)\b', r'\bfood\s+processing\b',
    r'\bretail\b', r'\becommerce\b', r'\bonline\s+shopping\b',
    r'\bfurniture\b', r'\btextile\b', r'\bapparel\b', r'\bfashion\b',
]

FUTURE_HEDGE = [
    r'\b(could|may|might|would|should|expect\w*|forecast\w*|predict\w*|anticipat\w*|project\w*)\b',
    r'\b(if|unless|in\s+case)\b',
    r'\b(risk\s+of|threat\s+of|fear\s+of|concern\s+about)\b',
    r'\blooming\b', r'\bpotential\b', r'\bpossible\b', r'\bpending\b',
]


def count_hits(text, patterns):
    return sum(1 for p in patterns if re.search(p, text, re.IGNORECASE))


def classify_v3(title: str, content: str) -> tuple[int, str]:
    txt_full = f"{title} {title} {title} {content}"
    title_l  = title.lower()
    full_l   = txt_full.lower()

    # Hard neutral
    neutral_hits = count_hits(full_l, NEUTRAL_HARD)
    if neutral_hits >= 2:
        return 0, f"NEUTRAL_HARD ({neutral_hits})"

    # Aerospace scope
    aerospace_hits = count_hits(full_l, AEROSPACE_RELATED)
    non_aero_hits  = count_hits(full_l, NON_AEROSPACE_HARD)
    apply_qt03     = (non_aero_hits >= 1 and aerospace_hits == 0)

    # HIGH score
    h_title = count_hits(title_l, ALL_HIGH_PATTERNS)
    h_full  = count_hits(full_l,  ALL_HIGH_PATTERNS)
    h_score = h_title * 3 + h_full

    # MEDIUM score
    m_title = count_hits(title_l, MEDIUM_PATTERNS)
    m_full  = count_hits(full_l,  MEDIUM_PATTERNS)
    m_score = m_title * 2 + m_full

    # Hedge (only used to downgrade HIGH->MEDIUM, NOT to block Label 1)
    hedge = count_hits(title_l + " " + content[:300], FUTURE_HEDGE)

    # ── Decision v3 ──────────────────────────────────────────────────────────
    if h_score >= 5 or h_title >= 2:
        raw = 2
        reason = f"HIGH: h_score={h_score} h_title={h_title}"
    elif h_score >= 3 and hedge <= 2:
        raw = 2
        reason = f"HIGH: h_score={h_score} hedge={hedge}"
    elif h_score >= 2 and m_score >= 2 and hedge <= 1:
        raw = 2
        reason = f"HIGH: h={h_score} m={m_score} hedge={hedge}"
    elif h_score >= 2 or (h_score == 1 and m_score >= 3):
        raw = 1
        reason = f"MEDIUM: h={h_score} m={m_score}"
    elif m_score >= 3:
        raw = 1
        reason = f"MEDIUM: m_score={m_score}"
    elif m_score >= 2 and neutral_hits == 0:
        raw = 1
        reason = f"MEDIUM: m_score={m_score} neutral=0"
    elif m_score >= 1:
        # v3: ha nguong -- bat ky tin hieu medium nao cung du -> Label 1
        raw = 1
        reason = f"MEDIUM: m_score={m_score} (v3 low-threshold)"
    else:
        raw = 0
        reason = f"NO_RISK: h={h_score} m={m_score}"

    # QT-03
    if apply_qt03 and raw > 0:
        old = raw
        raw = max(0, raw - 1)
        reason += f" | QT-03 ({old}->{raw})"

    return raw, reason


# ── Read Excel, classify, write back ─────────────────────────────────────────
print(f"Loading {XL_IN.name} ...")
wb = openpyxl.load_workbook(str(XL_IN))
ws = wb.worksheets[0]

LABEL_BG = {0: "E2EFDA", 1: "FFF2CC", 2: "FCE4D6"}
LABEL_FG = {0: "375623", 1: "7F6000", 2: "843C0C"}

def fill_cell(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

dist = {0: 0, 1: 0, 2: 0}
article_count = 0

for r in range(3, ws.max_row + 1):
    a_cell = ws.cell(r, 1)
    if a_cell.__class__.__name__ == "MergedCell":
        continue
    stt_val = a_cell.value
    if stt_val is None:
        continue
    try:
        stt = int(float(str(stt_val).strip()))
    except ValueError:
        continue

    title   = str(ws.cell(r, 5).value or "")
    content = str(ws.cell(r, 6).value or "")

    lbl, reason = classify_v3(title, content)
    dist[lbl] += 1
    article_count += 1

    c_g = ws.cell(r, 7)
    if c_g.__class__.__name__ == "MergedCell":
        continue
    c_g.value     = lbl
    c_g.font      = Font(bold=True, color=LABEL_FG[lbl], size=11, name="Calibri")
    c_g.fill      = fill_cell(LABEL_BG[lbl])
    c_g.alignment = Alignment(horizontal="center", vertical="center")
    c_g.border    = thin_border()

    if lbl > 0 or "QT-03" in reason:
        print(f"  [{stt:03d}] L={lbl} | {title[:75]}")
        print(f"         {reason}")

print(f"\nTotal: {article_count} articles")
print(f"Distribution v3: 0={dist[0]}  1={dist[1]}  2={dist[2]}")
print(f"Label 1 rate: {dist[1]/article_count*100:.1f}%  Label 2 rate: {dist[2]/article_count*100:.1f}%")

wb.save(str(XL_OUT))
print(f"\nSaved: {XL_OUT.name}")
