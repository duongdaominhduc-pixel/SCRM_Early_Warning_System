import openpyxl
import numpy as np
from statsmodels.stats.inter_rater import fleiss_kappa, aggregate_raters
from sklearn.metrics import cohen_kappa_score

wb = openpyxl.load_workbook("P0-03_pilot_labeling_150_v4.xlsx")
ws = wb.worksheets[0]

# Collect data rows (skip separator rows and rows with None labels)
data_rows = []
for r in range(3, ws.max_row + 1):
    g = ws.cell(r, 7).value
    h = ws.cell(r, 8).value
    i = ws.cell(r, 9).value
    j = ws.cell(r, 10).value
    if g is None or h is None or i is None or j is None:
        continue
    try:
        row = [int(float(g)), int(float(h)), int(float(i)), int(float(j))]
        data_rows.append(row)
    except (TypeError, ValueError):
        continue

all_data = np.array(data_rows)
block1 = all_data[:100]
block2 = all_data[100:]

print(f"Total valid articles: {len(all_data)} (Block1={len(block1)}, Block2={len(block2)})")

def fleiss(data, label=""):
    agg, cats = aggregate_raters(data)
    k = fleiss_kappa(agg)
    print(f"  {label}: kappa = {k:.4f}")
    return k

def level(k):
    if k >= 0.80: return "Almost Perfect"
    if k >= 0.70: return "Substantial"
    if k >= 0.60: return "Moderate"
    if k >= 0.40: return "Fair"
    if k >= 0.20: return "Slight"
    return "Poor"

rater_names = ["Claude", "Gemini", "DeepSeek", "ChatGPT"]

print("\n=== LABEL DISTRIBUTION ===")
dist = {}
for ri, name in enumerate(rater_names):
    col = all_data[:, ri]
    u, c = np.unique(col, return_counts=True)
    dist[name] = {int(u[k]): int(c[k]) for k in range(len(u))}
    print(f"  {name}: " + ", ".join(f"L{int(u[k])}={c[k]}" for k in range(len(u))))

print("\n=== FLEISS KAPPA ===")
k_all = fleiss(all_data, f"All {len(all_data)}")
k_b1  = fleiss(block1, f"Block1 ({len(block1)})")
k_b2  = fleiss(block2, f"Block2 ({len(block2)})")

print("\n=== PAIRWISE COHEN KAPPA ===")
pairs = [(0,1,"Claude","Gemini"),(0,2,"Claude","DeepSeek"),(0,3,"Claude","ChatGPT"),
         (1,2,"Gemini","DeepSeek"),(1,3,"Gemini","ChatGPT"),(2,3,"DeepSeek","ChatGPT")]
pair_results = []
for r1, r2, n1, n2 in pairs:
    k = cohen_kappa_score(all_data[:, r1], all_data[:, r2])
    pair_results.append((n1, n2, k))
    print(f"  {n1} vs {n2}: {k:.4f}  [{level(k)}]")

print("\n=== DISAGREEMENT ARTICLES (>=3 unique labels) ===")
disagreements = []
for idx, row in enumerate(all_data):
    if len(set(row)) >= 3:
        disagreements.append((idx+1, list(row)))
        print(f"  Article {idx+1:03d}: Claude={row[0]} Gemini={row[1]} DeepSeek={row[2]} ChatGPT={row[3]}")

if not disagreements:
    print("  (none)")

import json
results = {
    "total": len(all_data), "block1": len(block1), "block2": len(block2),
    "k_all": k_all, "k_b1": k_b1, "k_b2": k_b2,
    "pair_results": pair_results,
    "dist": dist,
    "disagreements": disagreements
}
with open("kappa_v2_results.json", "w", encoding="utf-8") as f:
    json.dump(results, f, indent=2)
print("\nSaved to kappa_v2_results.json")
