import openpyxl
import numpy as np
from statsmodels.stats.inter_rater import fleiss_kappa, aggregate_raters
from sklearn.metrics import cohen_kappa_score

wb = openpyxl.load_workbook("P0-03_pilot_labeling_150_v5.xlsx")
ws = wb.worksheets[0]

data_rows = []
for r in range(3, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a is None:
        continue
    try:
        stt = int(float(str(a).strip()))
    except ValueError:
        continue
    g = ws.cell(r, 7).value
    h = ws.cell(r, 8).value
    i = ws.cell(r, 9).value
    j = ws.cell(r, 10).value
    if None in (g, h, i, j):
        continue
    try:
        row = [int(float(g)), int(float(h)), int(float(i)), int(float(j))]
        data_rows.append((stt, row))
    except (TypeError, ValueError):
        continue

all_data = np.array([x[1] for x in data_rows])
block1 = all_data[:100]
block2 = all_data[100:]

names = ["Claude", "Gemini", "DeepSeek", "ChatGPT"]

def level(k):
    if k >= 0.80: return "Almost Perfect"
    if k >= 0.70: return "Substantial"
    if k >= 0.60: return "Moderate"
    if k >= 0.40: return "Fair"
    if k >= 0.20: return "Slight"
    return "Poor"

print(f"Valid articles: {len(all_data)} (Block1={len(block1)}, Block2={len(block2)})")

print("\n=== LABEL DISTRIBUTION ===")
dist = {}
for ri, name in enumerate(names):
    col = all_data[:, ri]
    u, c = np.unique(col, return_counts=True)
    dist[name] = {int(u[k]): int(c[k]) for k in range(len(u))}
    print(f"  {name}: " + ", ".join(f"L{int(u[k])}={c[k]}" for k in range(len(u))))

print("\n=== FLEISS KAPPA ===")
def fk(data, label):
    agg, _ = aggregate_raters(data)
    k = fleiss_kappa(agg)
    lv = level(k)
    print(f"  {label}: kappa={k:.4f}  [{lv}]")
    return k

k_all = fk(all_data, f"All {len(all_data)}")
k_b1  = fk(block1,   f"Block1 ({len(block1)})")
k_b2  = fk(block2,   f"Block2 ({len(block2)})")

print("\n=== PAIRWISE COHEN KAPPA ===")
pairs = [(0,1,"Claude","Gemini"),(0,2,"Claude","DeepSeek"),(0,3,"Claude","ChatGPT"),
         (1,2,"Gemini","DeepSeek"),(1,3,"Gemini","ChatGPT"),(2,3,"DeepSeek","ChatGPT")]
pair_results = []
for r1, r2, n1, n2 in pairs:
    k = cohen_kappa_score(all_data[:, r1], all_data[:, r2])
    pair_results.append((n1, n2, k))
    ok = "OK" if k >= 0.70 else ("~" if k >= 0.60 else "LOW")
    print(f"  {n1} vs {n2}: {k:.4f}  [{level(k)}]  {ok}")

print("\n=== DISAGREEMENT ARTICLES (>=3 unique labels) ===")
disagreements = []
for idx, (stt, row) in enumerate(data_rows):
    r = all_data[idx]
    if len(set(r)) >= 3:
        disagreements.append((stt, list(r)))
        print(f"  Art {stt:03d}: Claude={r[0]} Gemini={r[1]} DeepSeek={r[2]} ChatGPT={r[3]}")
if not disagreements:
    print("  (none)")

print(f"\nTotal disagreements: {len(disagreements)}")
print(f"\n{'='*50}")
print(f"FLEISS KAPPA = {k_all:.4f}  -->  {'PASS' if k_all >= 0.70 else 'FAIL'} (nguong 0.70)")
print(f"{'='*50}")
