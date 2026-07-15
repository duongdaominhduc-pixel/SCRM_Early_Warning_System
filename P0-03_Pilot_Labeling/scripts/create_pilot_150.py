"""
P0-03 — Pilot Labeling 150 bài (mở rộng từ 100)
Cấu trúc:
  - 100 bài: stratified random (giữ nguyên seed=42, KHÔNG thay đổi)
  - 50 bài bổ sung: purposive oversampling HIGH_RISK-heavy articles
    (lọc bài có keyword HIGH_RISK mạnh từ phần corpus còn lại)
  - Tổng: 150 bài, đảm bảo Nhan 2 >= 20 bai trong pilot set

Mục tiêu: per-category Kappa cho Nhan 2 co CI on dinh hon.
"""

import json, random, re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).parent.parent
NEWS = BASE / "0. news data" / "news_2022_2024_clean_2.json"
OUT  = Path(__file__).parent / "P0-03_pilot_labeling_150.xlsx"

SEED   = 42
STRATA = {"2022": 15, "2023": 38, "2024": 47}

random.seed(SEED)

# ── Load ──────────────────────────────────────────────────────────────────────
with open(NEWS, "r", encoding="utf-8") as f:
    all_articles = json.load(f)

by_year = defaultdict(list)
for art in all_articles:
    yr = str(art.get("publish_date", ""))[:4]
    if yr in STRATA:
        by_year[yr].append(art)

# ── Block 1: 100 bai stratified random (giong cu) ────────────────────────────
sampled_100 = []
for yr, n in STRATA.items():
    sampled_100.extend(random.sample(by_year[yr], n))
random.shuffle(sampled_100)
sampled_urls = {a["url"] for a in sampled_100}

# ── Classification logic (dung lai de label) ──────────────────────────────────
HIGH_RISK_PATTERNS = [
    r'\b(port|ports|terminal)\s+(shut\s*down|clos\w+|halt\w*|block\w*|suspend\w*)\b',
    r'\bshut\s*down\s+(port|terminal|plant|factory|facility)\b',
    r'\b(port|harbor|canal)\s+clos\w+\b',
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day|hit\s+|crippl\w+)\b',
    r'\b(workers?|dockworkers?|longshoremen|truckers?)\s+(on\s+strike|walk\w*\s+out|halted\s+work)\b',
    r'\bwork\s+stoppage\b', r'\bwalkout\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*|default\w*|collapse\w*)\b',
    r'\bfil\w+\s+for\s+bankruptcy\b',
    r'\b(disrupted|halted|suspended|blocked|stranded|grounded|paralyzed|crippled)\b',
    r'\b(flood|earthquake|typhoon|hurricane|wildfire|tsunami|cyclone)\s+\w*\s*(hit|struck|devastat\w*|destroy\w*|shut\w*|halt\w*)\b',
    r'\b(hit|struck|devastat\w*)\s+by\s+(flood|earthquake|typhoon|hurricane|wildfire|tsunami)\b',
    r'\b(sanction\w*|embargo|ban)\s+(imposed|effective|enforced|block\w*|tak\w+\s+effect)\b',
    r'\b(imposed|enforced)\s+sanction\w*\b',
    r'\b(plant|factory|facilit\w+|warehouse)\s+(shut\w*|clos\w*|halt\w*|idl\w+)\b',
    r'\b(shortage|shortfall|stockout|out\s+of\s+stock)\s+(of|in|for|hit\w*|confirm\w*|severe|critical|acute)\b',
    r'\b(severe|critical|acute)\s+(shortage|shortfall)\b',
    r'\b(vessel|ship|container|cargo)\s+(strand\w*|block\w*|detain\w*|seiz\w*)\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*)\b',
    r'\bsupplier\s+(fail\w*|bankrupt\w*|ceas\w+\s+operat\w*|shut\s+down)\b',
    r'\b(weeks?|months?)\s+of\s+delay\w*\b',
    r'\bdelay\w*\s+(of\s+)?(weeks?|months?)\b',
    r'\b(severe|critical|extreme|massive)\s+(port\s+congestion|backlog|bottleneck)\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?|containers?)\s+(wait\w*|strand\w*|anchor\w*)\b',
]
MEDIUM_RISK_PATTERNS = [
    r'\b(contract\s+talks?|labor\s+negotiat\w*|union\s+negotiat\w*|bargaining)\b',
    r'\b(strike\s+(threat|vote|authoriz\w*|warning|looming|loom\w*|possible|risk))\b',
    r'\b(workers?|union)\s+(vote|threaten|warn\w*|consider\w*)\s+\w*\s*(strike|walkout|action)\b',
    r'\bstrike\s+authoriz\w+\b',
    r'\b(increas\w+|grow\w+|ris\w+|worsening|escalat\w+)\s+(congestion|delay\w*|backlog|tension\w*|risk)\b',
    r'\b(congestion|delay\w*|backlog)\s+(increas\w+|grow\w+|ris\w+|worsening|escalat\w+)\b',
    r'\b(warn\w+|alert\w*|caution\w*|concern\w*)\s+(about|of|over|regard\w*)\s+\w*\s*(supply|chain|shortage|delay|disruption)\b',
    r'\bsupply\s+chain\s+(warn\w*|alert\w*|concern\w*|risk\w*|vulnerab\w*)\b',
    r'\b(propos\w+|plan\w*|consider\w*|draft\w*|introduc\w*)\s+(tariff|sanction\w*|restriction|ban)\b',
    r'\btariff\s+(hike|increase|propos\w*|threat\w*)\b',
    r'\btrade\s+(war|dispute|tension\w*|conflict)\b',
    r'\b(approaching|heading\s+toward|threaten\w*|expected\s+to\s+hit|forecast\w*)\b.{0,80}\b(port|ship\w*|supply|coast|region)\b',
    r'\b(storm|typhoon|hurricane|flood)\s+(threaten\w*|approach\w*|warn\w*|watch|warning)\b',
    r'\b(risk|threat|fear\w*|concern\w*)\s+(of\s+)?(shortage|shortfall|disruption|delay)\b',
    r'\bshortage\s+(risk|concern\w*|fear\w*|looming|possible|potential)\b',
    r'\bport\s+congestion\b',
    r'\bshipping\s+(delay\w*|backlog|bottleneck|disruption)\b',
    r'\bcargo\s+(delay\w*|backlog|stuck)\b',
    r'\bsupplier\s+(issue\w*|problem\w*|concern\w*|strain\w*|stress\w*|challeng\w*|struggl\w*)\b',
    r'\b(low|thin|tight|lean)\s+(inventory|stock|supply)\b',
    r'\binventory\s+(shortage|crunch|strain|pressure|concern)\b',
    r'\b(geopolitical|trade)\s+(tension\w*|uncertainty|instability|dispute\w*)\b',
    r'\b(longer|extended|stretched)\s+lead\s+time\w*\b',
    r'\blead\s+time\w*\s+(increase\w*|grow\w*|worsen\w*|stretch\w*|concern\w*)\b',
]
HIGH_BOOST_WORDS = [
    'closure','closed','shutdown','halted','suspended','stranded','bankruptcy',
    'bankrupt','defaulted','collapsed','strike action','work stoppage',
    'paralyzed','crippled','seized','blocked','detained',
]
MEDIUM_BOOST_WORDS = [
    'concern','warning','risk','potential','threat','looming',
    'negotiat','talks','pressure','tighten','escalat',
]
NOISE_PATTERNS_FOR_HIGH = [
    r'\b(forecast|outlook|predict\w*|expect\w*|project\w*|analys\w*)\b',
    r'\bhow\s+to\b', r'\bstrateg\w+\b', r'\binvest\w+\s+(in|for)\b',
]
STRONGLY_NEUTRAL = [
    r'\bartificial\s+intelligence\b', r'\bmachine\s+learning\b',
    r'\bdigital\s+(transform\w*|solution\w*)\b',
    r'\bwebinar\b', r'\bpodcast\b', r'\bwhitepaper\b',
    r'\bsurvey\s+(find\w*|show\w*|reveal\w*)\b',
    r'\bmarket\s+(share|size|growth|trend)\b',
]

def count_hits(text, patterns):
    t = text.lower()
    return sum(1 for p in patterns if re.search(p, t, re.IGNORECASE))

def classify(title, content):
    text_full  = f"{title} {title} {title} {content}"
    title_only = title.lower()
    neutral_hits = count_hits(text_full, STRONGLY_NEUTRAL)
    if neutral_hits >= 2:
        return 0
    h_pattern = count_hits(text_full, HIGH_RISK_PATTERNS)
    h_boost   = sum(1 for w in HIGH_BOOST_WORDS if w in text_full.lower())
    h_noise   = count_hits(content, NOISE_PATTERNS_FOR_HIGH)
    h_title   = count_hits(title_only, HIGH_RISK_PATTERNS)
    h_score   = h_pattern * 2 + h_boost + h_title * 3 - h_noise
    m_pattern = count_hits(text_full, MEDIUM_RISK_PATTERNS)
    m_boost   = sum(1 for w in MEDIUM_BOOST_WORDS if w in text_full.lower())
    m_title   = count_hits(title_only, MEDIUM_RISK_PATTERNS)
    m_score   = m_pattern + m_boost + m_title * 2
    if h_score >= 4 or h_title >= 2:       return 2
    if h_score >= 2 and m_score >= 2:      return 2
    if h_score >= 2 or (h_score == 1 and m_score >= 3): return 1
    if m_score >= 3:                        return 1
    if m_score >= 1 and h_score == 0 and neutral_hits == 0: return 1
    return 0

# ── HIGH_RISK scoring cho purposive sampling ──────────────────────────────────
def high_risk_score(art):
    text = f"{art.get('title','')} {art.get('title','')} {art.get('title','')} {art.get('content','')}"
    title = art.get('title','').lower()
    h = count_hits(text, HIGH_RISK_PATTERNS)
    hb = sum(1 for w in HIGH_BOOST_WORDS if w in text.lower())
    ht = count_hits(title, HIGH_RISK_PATTERNS)
    noise = count_hits(art.get('content',''), NOISE_PATTERNS_FOR_HIGH)
    return h * 2 + hb + ht * 3 - noise

# ── Block 2: 50 bai purposive HIGH_RISK oversampling ─────────────────────────
# Lay cac bai CHUA duoc chon, score cao -> sample 50 bai
pool_remaining = [a for a in all_articles
                  if str(a.get("publish_date",""))[:4] in STRATA
                  and a["url"] not in sampled_urls]

# Score tat ca bai con lai
scored = [(high_risk_score(a), a) for a in pool_remaining]
scored.sort(key=lambda x: -x[0])

# Lay top 200 bai co score cao nhat -> sample ngau nhien 50 de tranh bias
top_pool = [a for sc, a in scored if sc >= 3][:200]
random.seed(SEED + 1)   # seed khac de tranh trung lap pattern
if len(top_pool) >= 50:
    oversampled_50 = random.sample(top_pool, 50)
else:
    # Neu khong du 200 bai score>=3, lay top 50 truc tiep
    oversampled_50 = [a for _, a in scored[:50]]

print(f"Block 1 (random stratified): {len(sampled_100)} bai")
print(f"Block 2 (purposive HIGH_RISK): {len(oversampled_50)} bai")

# ── Gop va label toan bo 150 bai ─────────────────────────────────────────────
all_150 = []
for i, art in enumerate(sampled_100, 1):
    lbl = classify(art.get("title","") or "", art.get("content","") or "")
    all_150.append({"idx": i, "block": "RANDOM", "art": art, "label_claude": lbl})

for i, art in enumerate(oversampled_50, 1):
    lbl = classify(art.get("title","") or "", art.get("content","") or "")
    all_150.append({"idx": 100+i, "block": "HIGH_RISK_BOOST", "art": art, "label_claude": lbl})

dist_all = {0:0, 1:0, 2:0}
for row in all_150:
    dist_all[row["label_claude"]] += 1
print(f"\nDistribution 150 bai: {dist_all}")
print(f"  Label 2 rate: {dist_all[2]/150*100:.1f}%")

# ══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

LABEL_BG   = {0: "E2EFDA", 1: "FFF2CC", 2: "FCE4D6"}
LABEL_FG   = {0: "375623", 1: "7F6000", 2: "843C0C"}
BLOCK_BG   = {"RANDOM": "EBF3FB", "HIGH_RISK_BOOST": "FFF2CC"}

wb  = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: GAN NHAN
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "📋 Gán Nhãn"
ws1.freeze_panes = "E3"
ws1.sheet_view.zoomScale = 90

col_widths = {
    "A": 6, "B": 12, "C": 14, "D": 18,
    "E": 55, "F": 90,
    "G": 18, "H": 18, "I": 18, "J": 18,
    "K": 22,
}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# Banner row 1
ws1.merge_cells("A1:K1")
c = ws1["A1"]
c.value = ("PILOT LABELING — 150 BAI BAO  |  He thong EWS Chuoi Cung Ung  |  "
           "Nhap nhan: 0 / 1 / 2  (xem Sheet 'Huong Dan')  |  "
           "100 bai RANDOM + 50 bai HIGH_RISK BOOST")
c.font      = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
c.fill      = fill("1F3864")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 30

# Headers row 2
headers = [
    ("A2", "STT"),
    ("B2", "Article ID"),
    ("C2", "Ngay dang"),
    ("D2", "Nguon"),
    ("E2", "Tieu de bai bao"),
    ("F2", "Noi dung day du"),
    ("G2", "Nhan\nThanh vien 1\nDUC DAO - CLAUDE"),
    ("H2", "Nhan\nThanh vien 2\nTHUY - GEMINI"),
    ("I2", "Nhan\nThanh vien 3\nLINH - DEEPSEEK"),
    ("J2", "Nhan\nThanh vien 4\nHUYEN - CHAT GPT"),
    ("K2", "Ghi chu\n(kho / bat dong)"),
]
for addr, val in headers:
    c = ws1[addr]
    c.value     = val
    c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c.fill      = fill("2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
ws1.row_dimensions[2].height = 40

# Dropdown validation
dv = DataValidation(
    type="list", formula1='"0,1,2"', allow_blank=True,
    showInputMessage=True,
    promptTitle="Nhap nhan",
    prompt="0 = No Risk  |  1 = Medium Risk  |  2 = High Risk",
    showErrorMessage=True,
    errorTitle="Nhan khong hop le",
    error="Chi nhap 0, 1, hoac 2",
)
ws1.add_data_validation(dv)
dv.sqref = "G3:J152"

# Data rows
for row_data in all_150:
    i    = row_data["idx"]
    row  = i + 2
    art  = row_data["art"]
    lbl  = row_data["label_claude"]
    blk  = row_data["block"]

    base_fill = fill("EBF3FB" if i % 2 == 1 else "FFFFFF")
    if blk == "HIGH_RISK_BOOST":
        base_fill = fill("FFF9EC" if i % 2 == 1 else "FFFDF5")

    cells = {
        f"A{row}": (i,                          "center", base_fill),
        f"B{row}": (f"ART-{i:03d}",             "center", base_fill),
        f"C{row}": (art.get("publish_date",""), "center", base_fill),
        f"D{row}": (art.get("source",""),        "center", base_fill),
        f"E{row}": (art.get("title",""),         "left",   base_fill),
        f"F{row}": (art.get("content","") or "", "left",   base_fill),
        f"G{row}": (lbl,                         "center", fill(LABEL_BG[lbl])),
        f"H{row}": ("",                          "center", fill("FFF2CC")),
        f"I{row}": ("",                          "center", fill("FFF2CC")),
        f"J{row}": ("",                          "center", fill("FFF2CC")),
        f"K{row}": ("",                          "left",   base_fill),
    }
    for addr, (val, align, bg) in cells.items():
        c           = ws1[addr]
        c.value     = val
        c.fill      = bg
        c.border    = thin_border()
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        if addr[0] == "G" and val != "":
            c.font = Font(bold=True, color=LABEL_FG[lbl], size=11, name="Calibri")
        else:
            c.font = Font(size=9, name="Calibri")

    ws1.row_dimensions[row].height = 90

# Separator row between block 1 và block 2
sep_row = 103
ws1.insert_rows(sep_row)
ws1.merge_cells(f"A{sep_row}:K{sep_row}")
c = ws1[f"A{sep_row}"]
c.value     = ("--- 50 BAI BO SUNG (HIGH_RISK BOOST) ---  "
               "Cac bai nay duoc chon co chu dich de tang so luong Nhan 2. "
               "Van gan nhan binh thuong theo tieu chi.")
c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
c.fill      = fill("843C0C")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[sep_row].height = 22

# Footer
footer_row = 154
ws1.merge_cells(f"A{footer_row}:K{footer_row}")
c = ws1[f"A{footer_row}"]
c.value = (f"TONG: 150 BAI  |  Block 1 (Random Stratified): 100 bai  "
           f"|  Block 2 (High Risk Boost): 50 bai  "
           f"|  Claude label dist: 0={dist_all[0]}  1={dist_all[1]}  2={dist_all[2]}")
c.font      = Font(bold=True, color="1F3864", size=9, name="Calibri")
c.fill      = fill("D9E1F2")
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[footer_row].height = 18

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: HUONG DAN (giu nguyen logic cu, chi update so bai)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("📖 Hướng Dẫn")
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 100
ws2.sheet_view.zoomScale = 95

guide = [
    (None, "1F3864","FFFFFF",True,14,
     "HUONG DAN GAN NHAN — PILOT LABELING 150 BAI",35),
    (None,"2E75B6","FFFFFF",True,11,"CAU TRUC 150 BAI",22),
    (None,None,"000000",False,10,
     "Block 1 — 100 bai RANDOM STRATIFIED (bai 001-100):\n"
     "  Lay ngau nhien tu toan bo corpus, dam bao dai dien theo nam (2022/2023/2024).\n"
     "  Nen co: nhieu Nhan 0, mot so Nhan 1, it Nhan 2.\n\n"
     "Block 2 — 50 bai HIGH_RISK BOOST (bai 101-150):\n"
     "  Chon co chu dich tu cac bai co tu khoa HIGH_RISK manh.\n"
     "  Muc dich: tang so luong Nhan 2 len >= 20 bai de Kappa per-category on dinh.\n"
     "  LUU Y: day la oversampling co chu dich — KHONG co nghia la tat ca deu la Nhan 2!\n"
     "  Van gan nhan binh thuong theo tieu chi, co the la 0/1/2 tuy noi dung thuc te.",90),
    (None,"2E75B6","FFFFFF",True,11,"BANG NHAN",22),
    ("A","E2EFDA","375623",True,10,"Nhan 0 — NO RISK",22),
    ("B","E2EFDA","375623",False,10,
     "Bai thong tin chung, phan tich thi truong, khong co su kien gian doan cu the.\n"
     "Vi du: 'Freight rates expected to stabilize in Q2'",45),
    ("A","FFF2CC","7F6000",True,10,"Nhan 1 — MEDIUM RISK",22),
    ("B","FFF2CC","7F6000",False,10,
     "Co dau hieu gian doan DANG XUAT HIEN hoac CO THE xay ra.\n"
     "Bao gom: tac nghen cang tang nhe, tranh chap lao dong dang thuong luong, thien tai chua anh huong truc tiep.\n"
     "Vi du: 'Union workers vote on strike authorization'",60),
    ("A","FCE4D6","843C0C",True,10,"Nhan 2 — HIGH RISK",22),
    ("B","FCE4D6","843C0C",False,10,
     "Su kien gian doan DA XAY RA hoac CO TAC DONG NGAY LAP TUC.\n"
     "Bao gom: cang dong cua, dinh cong thuc te, pha san nha cung cap, lenh trung phat hieu luc.\n"
     "Vi du: 'Port of LA shuts down — 40 vessels stranded'",65),
    (None,"2E75B6","FFFFFF",True,11,"TIEU CHI GAN NHAN",22),
    (None,None,"000000",False,10,
     "Doc KY tieu de VA noi dung truoc khi gan.\n"
     "Tap trung vao TAC DONG den chuoi cung ung, khong phai muc do nghiem trong cua su kien.\n"
     "Neu khong chac chan giua 2 nhan → gan nhan THAP hon va ghi chu vao cot K.\n"
     "KHONG gan Nhan 2 cho cac bai chi de cap rui ro tiem an ma khong co su kien thuc te.",75),
    (None,"2E75B6","FFFFFF",True,11,"PHAN CONG",22),
    (None,None,"000000",False,10,
     "Cot G  →  Thanh vien 1: DUC DAO - CLAUDE (DA DIEN SAN)\n"
     "Cot H  →  Thanh vien 2: THUY - GEMINI\n"
     "Cot I   →  Thanh vien 3: LINH - DEEPSEEK\n"
     "Cot J  →  Thanh vien 4: HUYEN - CHAT GPT\n\n"
     "Deadline: ___________________",75),
]

r = 1
for row_data in guide:
    merge_end, bg, fg, bold, size, text, height = row_data
    ws2.row_dimensions[r].height = height
    if merge_end is None:
        ws2.merge_cells(f"A{r}:B{r}")
        c = ws2[f"A{r}"]
    else:
        c = ws2[f"B{r}"]
        a = ws2[f"A{r}"]
        a.value = ""
        if bg: a.fill = fill(bg)
        a.border = thin_border()
    c.value     = text
    c.font      = Font(bold=bold, color=fg or "000000", size=size, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if bg: c.fill = fill(bg)
    c.border = thin_border()
    r += 1

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3: FLEISS KAPPA SCRIPT
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("🧮 Fleiss Kappa")
ws3.column_dimensions["A"].width = 110
ws3.sheet_view.zoomScale = 90

import textwrap
kappa_script = textwrap.dedent('''\
"""
P0-03: Tinh Fleiss Kappa tu file Excel 150 bai da hoan thanh.
Chay script nay SAU KHI ca 4 thanh vien dien nhan vao cot G, H, I, J.
"""

import pandas as pd
import numpy as np
import json
from pathlib import Path
from statsmodels.stats.inter_rater import fleiss_kappa, aggregate_raters

EXCEL_PATH = Path(__file__).parent / "P0-03_pilot_labeling_150.xlsx"

def run_fleiss_kappa():
    df = pd.read_excel(
        EXCEL_PATH,
        sheet_name="📋 Gán Nhãn",
        header=1,          # Row 2 la header
        usecols="G:J",     # 4 cot nhan
        nrows=152,         # doc du de cover 150 bai + separator row
    )
    df.columns = ["Claude", "Gemini", "DeepSeek", "ChatGPT"]

    # Bo qua separator row (row co gia tri NaN hoac text)
    df = df[pd.to_numeric(df["Claude"], errors="coerce").notna()]
    df = df.apply(pd.to_numeric, errors="coerce").dropna()
    df = df.astype(int)

    n_items = len(df)
    print(f"=== DU LIEU ===")
    print(f"So bai da gan nhan du (ca 4 raters): {n_items} / 150")

    if n_items < 150:
        missing = 150 - n_items
        print(f"CANH BAO: Con {missing} bai chua du nhan. Hoan thanh truoc khi tinh Kappa.")
        if n_items < 50:
            print("Qua it du lieu, dung lai.")
            return

    labels = df.values   # shape (n_items, 4), values 0/1/2

    # Phan phoi nhan theo tung rater
    print("\\n=== PHAN PHOI NHAN ===")
    for i, rater in enumerate(["Claude (TV1)", "Gemini (TV2)", "DeepSeek (TV3)", "ChatGPT (TV4)"]):
        v = labels[:, i]
        print(f"  {rater}: 0={np.sum(v==0):3d}  1={np.sum(v==1):3d}  2={np.sum(v==2):3d}")

    # Fleiss Kappa toan bo
    agg, cats = aggregate_raters(labels)
    kappa_overall = fleiss_kappa(agg, method="fleiss")

    print(f"\\n=== FLEISS KAPPA ===")
    print(f"  Kappa (150 bai): {kappa_overall:.4f}")

    # Kappa rieng Block 1 (100 bai random) va Block 2 (50 bai boost)
    if n_items >= 100:
        agg1, _ = aggregate_raters(labels[:100])
        k1 = fleiss_kappa(agg1, method="fleiss")
        print(f"  Kappa Block 1 (100 bai random): {k1:.4f}")
    if n_items >= 150:
        agg2, _ = aggregate_raters(labels[100:150])
        k2 = fleiss_kappa(agg2, method="fleiss")
        print(f"  Kappa Block 2 (50 bai HR boost): {k2:.4f}")

    # Phien giai
    k = kappa_overall
    if   k >= 0.80: level, status = "Almost Perfect",  "PASS"
    elif k >= 0.70: level, status = "Substantial",     "PASS"
    elif k >= 0.60: level, status = "Moderate",        "CANH BAO - chua dat 0.70"
    elif k >= 0.40: level, status = "Fair",             "FAIL - can hop lai"
    else:           level, status = "Poor/Slight",      "FAIL NGHIEM TRONG"

    print(f"  Muc do: {level}")
    print(f"  Danh gia: {status}")
    print(f"  Nguong yeu cau: kappa >= 0.70")

    # Bat dong
    disagreement = []
    for idx in range(len(labels)):
        row_l = labels[idx]
        if len(set(row_l)) >= 3:
            disagreement.append({
                "STT": idx + 1,
                "Article_ID": f"ART-{idx+1:03d}",
                "Labels": row_l.tolist(),
            })
    print(f"\\n=== BAT DONG ===")
    print(f"  So bai bat dong (>=3 nhan khac nhau): {len(disagreement)}")
    for d in sorted(disagreement, key=lambda x: -len(set(x["Labels"])))[:10]:
        print(f"    {d[\'Article_ID\']}: {d[\'Labels\']}")

    result = {
        "fleiss_kappa_overall": round(k, 4),
        "level": level,
        "status": "PASS" if k >= 0.70 else "FAIL",
        "n_items": n_items,
        "n_disagreement": len(disagreement),
        "disagreement_articles": disagreement,
    }
    out = Path(__file__).parent / "fleiss_kappa_result.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\\n-> Ket qua luu: {out}")
    return k

if __name__ == "__main__":
    run_fleiss_kappa()
''')

kappa_sections = [
    ("1F3864","FFFFFF",True,14,"FLEISS KAPPA — Script Tinh Do Dong Thuan (150 bai / 4 Raters)",35),
    ("2E75B6","FFFFFF",True,11,"BUOC 1 — CAI DAT (chay 1 lan trong terminal)",24),
    ("1E1E1E","00FF7F",False,9,"pip install statsmodels openpyxl pandas",22),
    ("2E75B6","FFFFFF",True,11,"BUOC 2 — SCRIPT CHINH (copy vao compute_fleiss_kappa.py)",24),
    ("1E1E1E","D4D4D4",False,9, kappa_script, 1200),
    ("2E75B6","FFFFFF",True,11,"BUOC 3 — PHIEN GIAI KET QUA",24),
    ("E2EFDA","375623",False,10,
     "Thang do Landis & Koch (1977):\n\n"
     "  kappa < 0.00  → Kem (Poor)\n"
     "  0.00 – 0.20   → Khong dang ke (Slight)\n"
     "  0.21 – 0.40   → Trung binh (Fair)\n"
     "  0.41 – 0.60   → Vua phai (Moderate)\n"
     "  0.61 – 0.80   → Dang ke (Substantial)   <- MUC TIEU: kappa >= 0.70\n"
     "  0.81 – 1.00   → Gan hoan hao (Almost Perfect)\n\n"
     "GATE: kappa >= 0.70 → tien hanh pseudo-labeling 700 bai\n"
     "Neu FAIL → DUNG, bao cao truong nhom de dieu chinh guideline",160),
    ("FCE4D6","843C0C",False,10,
     "XU LY KHI KAPPA THAP:\n\n"
     "1. Xem danh sach bai bat dong (disagreement_articles trong JSON output)\n"
     "2. Hop nhom, doc lai bai bat dong, thao luan tai sao nhan khac nhau\n"
     "3. Cap nhat vi du trong Huong Dan\n"
     "4. Neu van < 0.60 → xem xet merge Nhan 1+2 thanh AT_RISK (binary)\n"
     "5. Bao cao truong nhom truoc khi quyet dinh",120),
]

r = 1
for bg, fg, bold, size, text, height in kappa_sections:
    ws3.row_dimensions[r].height = height
    c = ws3[f"A{r}"]
    c.value     = text
    c.font      = Font(bold=bold, color=fg, size=size,
                       name="Courier New" if bg=="1E1E1E" else "Calibri")
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border    = thin_border()
    r += 1

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Sheets: Gan Nhan (150 bai) | Huong Dan | Fleiss Kappa")
