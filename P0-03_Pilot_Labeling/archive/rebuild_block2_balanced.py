"""
P0-03 — Rebuild Block 2 BALANCED: 25 bai Nhan 2 ro rang + 25 bai Nhan 1 ro rang
Muc tieu phan phoi 150 bai:
  Nhan 0: ~70 bai (Block 1 random chu yeu)
  Nhan 1: ~35 bai (9 tu Block 1 + 25 title-filtered MEDIUM)
  Nhan 2: ~45 bai (20 tu Block 1 + 25 title-filtered HIGH)
"""

import json, random, re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = Path(__file__).parent.parent
NEWS = BASE / "0. news data" / "news_2022_2024_clean_2.json"
XL   = Path(__file__).parent / "P0-03_pilot_labeling_150_v4.xlsx"

SEED = 42; STRATA = {"2022": 15, "2023": 38, "2024": 47}

random.seed(SEED)
with open(NEWS, "r", encoding="utf-8") as f:
    all_articles = json.load(f)

by_year = defaultdict(list)
for art in all_articles:
    yr = str(art.get("publish_date", ""))[:4]
    if yr in STRATA:
        by_year[yr].append(art)

# Block 1: giu nguyen 100 bai random
sampled_100 = []
for yr, n in STRATA.items():
    sampled_100.extend(random.sample(by_year[yr], n))
random.shuffle(sampled_100)
used_urls = {a["url"] for a in sampled_100}

print(f"Block 1 giu nguyen: {len(sampled_100)} bai")

# ══════════════════════════════════════════════════════════════════════════════
# TITLE FILTERS
# ══════════════════════════════════════════════════════════════════════════════

# --- HIGH_RISK title patterns (su kien DA xay ra) ----------------------------
TITLE_HIGH = [
    # Bankruptcy
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\bfil\w+\s+for\s+(bankruptcy|chapter\s+11)\b',
    r'\b(goes?|went|filed)\s+(bankrupt|into\s+administration)\b',
    r'\bdefault\w*\s+on\s+(debt|payment|bond|loan)\b',
    r'\bceas\w+\s+(operations?|trading|business)\b',
    # Strike ACTIVE
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day|launched|halts?|shuts?|crippl\w+)\b',
    r'\b(workers?|dockworkers?|longshoremen|port\s+workers?|rail\s+workers?)\s+(on\s+strike|walk\w+\s+out|began\s+strik\w+)\b',
    r'\bwalkout\s+(begins|started|underway|hits?|halts?)\b',
    r'\bwork\s+stoppage\s+(begins|at|hits?|halts?)\b',
    r'\bpicket\s+(line|lines)\s+(block\w*|halt\w*)\b',
    # Port/plant shutdown
    r'\b(port|terminal|harbor)\s+(shut\s*down|closure|closed|halted|suspended)\b',
    r'\bport\s+of\s+\w[\w\s]{0,20}(shut\w*|clos\w*|halt\w*)\b',
    r'\b(plant|factory|facilit\w+)\s+(shut\w*|clos\w*|halt\w*|idl\w+|suspended)\b',
    r'\bproduction\s+(halted|suspended|shut\s+down|stopped)\b',
    # Vessel/cargo seized
    r'\b(ship|vessel|cargo|tanker)\s+(seized|detained|impounded|blocked|stranded)\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*)\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?)\s+(strand\w*|block\w*|wait\w*)\b',
    # Houthi / Red Sea active
    r'\bhouthi\w*\s+(attack\w*|fire\w*|hit\s+|struck|seized|hijack\w*)\b',
    r'\bcrew\s+(evacuated|rescued)\b.{0,40}\b(red\s+sea|houthi|attack)\b',
    r'\b(maersk|hapag|msc|cma\s*cgm|evergreen|cosco|zim)\s+\w{0,15}\s*(suspend\w*|halt\w*|divert\w*|reroute\w*)\b',
    # Lockdown active
    r'\blockdown\s+(hits?|halt\w*|shut\w*|crippl\w*|jeopardiz\w*|disrupt\w*)\b',
    r'\b(covid|covid-19)\s+(lockdown|shutdown|closure)\b',
    r'\bzero[- ]covid\s+(lockdown|shutdown)\b',
    # Disaster confirmed impact
    r'\b(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami)\s+\w{0,20}\s*(halt\w*|shut\w*|clos\w*|disrupt\w*|crippl\w*|destroy\w*)\b',
    r'\b(halt\w*|shut\w*|clos\w*|disrupt\w*)\s+\w{0,20}\s*(hurricane|typhoon|flood|earthquake|wildfire)\b',
    # Shortage confirmed
    r'\b(stockout|out\s+of\s+stock|ran\s+out\s+of)\b',
    r'\b(critical|acute|severe)\s+shortage\b',
    r'\bsupply\s+(crisis|collapse|failure)\s+(hit\w*|crippl\w*)\b',
]

# --- MEDIUM_RISK title patterns (dau hieu DANG hinh thanh) -------------------
TITLE_MEDIUM = [
    # Strike PENDING: dam phan, bieu quyet, canh bao
    r'\bstrike\s+(threat\w*|vote|authoriz\w*|warning|loom\w*|possible|risk|could|feared|liken\w*)\b',
    r'\b(workers?|union|longshoremen|dockworkers?|employees?)\s+(vote|threaten\w*|warn\w*|prepar\w*|consider\w*|reject\w*)\s+\w{0,20}\s*(strike|walkout|action)\b',
    r'\bstrike\s+authoriz\w+\b',
    r'\b(contract\s+talks?|labor\s+negotiat\w*|union\s+negotiat\w*|collective\s+bargaining)\b',
    r'\bcontract\s+(expir\w*|deadline|impasse|breakdown|reject\w*)\b',
    r'\bwork\s+(action|dispute)\s+(loom\w*|threat\w*|possible)\b',
    r'\b(rail|railway|port|dock)\s+workers?\s+(threaten\w*|warn\w*|plan\w*|vote)\b',
    # Red Sea / Houthi: canh bao chua tan cong
    r'\b(warn\w*|alert\w*|caution\w*|advis\w*)\b.{0,50}\b(red\s+sea|houthi|shipping|route)\b',
    r'\bhouthi\w*\s+(threat\w*|warn\w*|target\w*|could|may)\b',
    r'\bshipping\s+(risk|warn\w*|alert\w*|concern\w*)\b.{0,40}\b(red\s+sea|gulf|Yemen)\b',
    r'\binsurance\s+(surcharge|premium|rate)\w*\s+(rise\w*|increas\w*|surge\w*|soar\w*)\b',
    # Tariff / sanction: dang de xuat
    r'\b(propos\w+|plan\w*|consider\w*|mull\w*|weigh\w*|draft\w*)\s+(tariff|sanction\w*|restriction|ban|levy)\b',
    r'\btariff\s+(hike|increase|propos\w*|threat\w*|possible|could|plan\w*)\b',
    r'\btrade\s+(war|dispute|tension\w*|conflict|friction|row)\b',
    r'\b(us|eu|china|uk)\s+(consider\w*|plan\w*|weigh\w*|discuss\w*|threaten\w*)\s+(sanction\w*|ban|tariff|duty)\b',
    r'\bnew\s+(tariff|duty|sanction)\s+(plan\w*|propos\w*|consider\w*|possible)\b',
    # Disaster: dang tien den, chua do bo
    r'\b(storm|typhoon|hurricane|flood|cyclone)\s+(approach\w*|threaten\w*|head\w+\s+toward|forecast\w*|warn\w*|watch|expected\s+to\s+hit|bearing\s+down)\b',
    r'\b(category\s+\d|tropical\s+storm)\s+\w+\s+(approach\w*|threaten\w*|target\w*|aims?\s+at)\b',
    # Supplier distress / geopolitical
    r'\bsupplier\s+(distress\w*|strain\w*|warn\w*|concern\w*|struggl\w*|at\s+risk)\b',
    r'\b(financial\s+distress|cash\s+crunch|liquidity\s+crisis|credit\s+crunch)\b',
    r'\bgeopolitical\s+(tension\w*|uncertainty|instability|risk\w*)\b',
    r'\bsupply\s+chain\s+(risk\w*|vulnerab\w*|concern\w*|warn\w*|threat\w*|uncertain\w*)\b',
    # Port congestion rising / shipping disruption building
    r'\bport\s+congestion\s+(grow\w*|worsen\w*|increas\w*|rise\w*|build\w*)\b',
    r'\bshipping\s+(disruption|backlog|bottleneck)\s+(grow\w*|worsen\w*|rise\w*|build\w*|loom\w*)\b',
    r'\bcargo\s+(backlog|delay|stuck)\b',
    r'\blead\s+time\w*\s+(increase\w*|grow\w*|stretch\w*|lengthen\w*)\b',
    # Shortage risk building
    r'\bshortage\s+(risk|concern\w*|fear\w*|loom\w*|possible|potential|ahead|warn\w*)\b',
    r'\b(tight|thin|low|lean)\s+(supply|inventory|stock)\b',
    r'\binventory\s+(strain|pressure|concern|crunch|depletion)\b',
    r'\b(risk|threat|fear\w*|concern\w*)\s+(of\s+)?(shortage|shortfall|disruption|supply\s+crunch)\b',
    # Panama Canal low water (restriction, chua dong)
    r'\bpanama\s+canal\s+(restrict\w*|drought|low\s+water|limit\w*|backlog|jam)\b',
    # Airline/carrier financial stress (chua pha san)
    r'\b(airline|carrier|shipper)\s+(financial\s+)?stress\w*\b',
    r'\bdebt\s+(restructur\w*|concern\w*|worry|negoti\w*)\b',
    r'\bcreditor\w*\s+(negoti\w*|deal|plan\w*)\b',
]

# --- Filter: KHONG DUOC co cac tu khoa phu dinh de tranh nhap -------------------
# Bai Medium KHONG duoc la Nhan 2 (khong co event DA xay ra)
EXCLUDE_FROM_MEDIUM = [
    r'\bstrike\s+(begins|started|underway|enters\s+day|launched)\b',
    r'\b(workers?|dockworkers?)\s+(on\s+strike|walk\w+\s+out)\b',
    r'\bwalkout\s+(begins|started|underway)\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*)\b',
    r'\b(port|terminal)\s+(shut\s*down|closure|closed|halted)\b',
    r'\bhouthi\w*\s+(attack\w*|fire\w*|hit|struck|seized)\b',
    r'\blockdown\s+(hits?|halt\w*|shut\w*|crippl\w*)\b',
    r'\bproduction\s+(halted|suspended|shut\s+down)\b',
]


def title_high_score(title):
    t = title.lower()
    return sum(1 for p in TITLE_HIGH if re.search(p, t, re.I))


def title_medium_score(title):
    t = title.lower()
    m = sum(1 for p in TITLE_MEDIUM if re.search(p, t, re.I))
    excl = sum(1 for p in EXCLUDE_FROM_MEDIUM if re.search(p, t, re.I))
    return m if excl == 0 else 0


# ── Score tat ca bai con lai ──────────────────────────────────────────────────
pool = [a for a in all_articles
        if str(a.get("publish_date", ""))[:4] in STRATA
        and a["url"] not in used_urls]

high_pool, medium_pool = [], []
for art in pool:
    t = art.get("title", "") or ""
    hs = title_high_score(t)
    ms = title_medium_score(t)
    if hs >= 1:
        high_pool.append((hs, art))
    elif ms >= 1:
        medium_pool.append((ms, art))

high_pool.sort(key=lambda x: -x[0])
medium_pool.sort(key=lambda x: -x[0])

print(f"\nPool HIGH  (title score>=1): {len(high_pool)} bai")
print(f"Pool MEDIUM (title score>=1): {len(medium_pool)} bai")

# Sample 25 HIGH tu top 80
random.seed(SEED + 2)
top_high = [a for _, a in high_pool[:80]]
block2_high = random.sample(top_high, min(25, len(top_high)))

# Sample 25 MEDIUM tu top 80
top_medium = [a for _, a in medium_pool[:80]]
block2_medium = random.sample(top_medium, min(25, len(top_medium)))

block2_new = block2_high + block2_medium
random.shuffle(block2_new)

print(f"\nBlock 2 moi: {len(block2_high)} HIGH + {len(block2_medium)} MEDIUM = {len(block2_new)} bai")

print("\n--- 25 bai HIGH_RISK (da xay ra) ---")
for i, a in enumerate(block2_high, 1):
    print(f"  H{i:02d} | {a.get('title','')[:85]}")
print("\n--- 25 bai MEDIUM_RISK (dang hinh thanh) ---")
for i, a in enumerate(block2_medium, 1):
    print(f"  M{i:02d} | {a.get('title','')[:85]}")

# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFIER v2
# ══════════════════════════════════════════════════════════════════════════════
ALL_HIGH_CONTENT = [
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day\s*\d|launched)\b',
    r'\b(workers?|dockworkers?|longshoremen)\s+(on\s+strike|walk\w+\s+out|began\s+strik\w+)\b',
    r'\bwalkout\s+(begins|started|underway)\b', r'\bpicket\s+line\b',
    r'\bhouthi\w*\s+(attack\w*|fire\w*|hit\s+|struck|seized)\b',
    r'\b(maersk|hapag|msc|cma\s*cgm|evergreen)\s+\w*\s*(suspend\w*|halt\w*|divert\w*|reroute\w*)\b',
    r'\b(port|terminal)\s+(is\s+)?(clos\w+|shut\s*down|halt\w*|suspend\w*)\b',
    r'\b(plant|factory)\s+(shut\w*|clos\w*|halt\w*|idl\w+)\b',
    r'\bproduction\s+(halt\w*|suspend\w*|shut\w*)\b',
    r'\bfil\w+\s+for\s+(bankruptcy|chapter\s+11)\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\b(hurricane|typhoon|flood\w*|earthquake)\s+\w*\s*(hit\w*|struck|devastat\w*|halt\w*)\b',
    r'\b(sanction\w*|embargo)\s+(in\s+effect|effective|enforced|imposed)\b',
    r'\b(lockdown|lock\s+down)\s+(in\s+effect|imposed|begins|started)\b',
    r'\b(covid|covid-19)\s+(lockdown|shutdown)\b',
    r'\b(stockout|out\s+of\s+stock|critical\s+shortage|acute\s+shortage)\b',
    r'\b(vessel|ship|cargo)\s+(seized|detained|blocked|stranded)\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*)\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?)\s+(strand\w*|wait\w*)\b',
    r'\bdelays?\s+of\s+(several\s+)?(weeks?|months?)\b',
]
MEDIUM_CONTENT = [
    r'\b(contract\s+talks?|labor\s+negotiat\w*|union\s+negotiat\w*)\b',
    r'\bstrike\s+(threat|vote|authoriz\w*|warning|loom\w*|possible)\b',
    r'\b(workers?|union)\s+(vote|threaten|warn\w*)\s+\w*\s*(strike|walkout)\b',
    r'\bport\s+congestion\b',
    r'\bshipping\s+(delay\w*|backlog|bottleneck|disruption)\b',
    r'\b(risk|threat|concern\w*)\s+(of\s+)?(shortage|shortfall|disruption)\b',
    r'\bsupplier\s+(issue\w*|strain\w*|distress\w*|concern\w*)\b',
    r'\btrade\s+(war|dispute|tension\w*)\b',
    r'\btariff\s+(hike|propos\w*|threat\w*)\b',
    r'\b(storm|typhoon|hurricane)\s+(threaten\w*|approach\w*|warn\w*|watch)\b',
    r'\b(financial\s+distress|cash\s+crunch)\b',
    r'\bgeopolitical\s+(tension\w*|uncertainty|instability)\b',
    r'\bsupply\s+chain\s+(risk\w*|concern\w*|warn\w*|vulnerab\w*)\b',
    r'\b(tight|thin|lean)\s+(supply|inventory|stock)\b',
    r'\bshortage\s+(risk|concern\w*|loom\w*|possible)\b',
    r'\blead\s+time\w*\s+(increase\w*|grow\w*|stretch\w*)\b',
]
NEUTRAL_HARD = [
    r'\bartificial\s+intelligence\b', r'\bmachine\s+learning\b',
    r'\bdigital\s+(transform\w*|platform\w*)\b', r'\bwebinar\b',
    r'\baward\w*\s+(winner|recipient)\b', r'\bnew\s+(ceo|cfo|vp|director)\b',
    r'\bmarket\s+(share|size|growth|forecast)\b',
    r'\bprofit\s+(rise\w*|increas\w*|beat\w*)\b', r'\bipo\b',
]
AEROSPACE_RELATED = [
    r'\baerospace\b', r'\baviation\b', r'\baircraft\b', r'\bairline\b',
    r'\bair\s+(cargo|freight)\b', r'\bboeing\b', r'\bairbus\b',
    r'\b(port|shipping|freight|container|suez|red\s+sea|panama)\b',
    r'\b(semiconductor|energy\s+supply|steel|copper|rare\s+earth)\b',
]
NON_AEROSPACE_HARD = [
    r'\b(automotive|auto\s+(industry|maker|plant))\b',
    r'\belectric\s+vehicle\b', r'\bev\s+(battery|maker)\b',
    r'\b(farm\w*|agricultur\w*|grain|wheat|soybean|fertilizer)\b',
    r'\b(pharmaceutical|drug\s+(supply|shortage))\b',
    r'\b(food\s+(supply|chain|shortage)|food\s+processing)\b',
    r'\b(retail|ecommerce|fashion|apparel|textile)\b',
]
FUTURE_HEDGE = [
    r'\b(could|may|might|expect\w*|forecast\w*|anticipat\w*)\b',
    r'\b(risk\s+of|threat\s+of|fear\s+of)\b',
    r'\b(looming|potential|possible|pending)\b',
]


def cnt(text, pats):
    t = text.lower()
    return sum(1 for p in pats if re.search(p, t, re.I))


def classify_v2(title, content):
    full = f"{title} {title} {title} {content}"
    tl = title.lower()
    neutral = cnt(full, NEUTRAL_HARD)
    if neutral >= 2:
        return 0
    aero     = cnt(full, AEROSPACE_RELATED)
    non_aero = cnt(full, NON_AEROSPACE_HARD)
    qt03     = (non_aero >= 1 and aero == 0)
    h_t = cnt(tl, ALL_HIGH_CONTENT)
    h_f = cnt(full, ALL_HIGH_CONTENT)
    h   = h_t * 3 + h_f
    m_t = cnt(tl, MEDIUM_CONTENT)
    m_f = cnt(full, MEDIUM_CONTENT)
    m   = m_t * 2 + m_f
    hedge = cnt(tl + " " + content[:500], FUTURE_HEDGE)

    if h >= 5 or h_t >= 2:              raw = 2
    elif h >= 3 and hedge <= 2:         raw = 2
    elif h >= 2 and m >= 2 and hedge <= 1: raw = 2
    elif h >= 2 or (h == 1 and m >= 4): raw = 1
    elif m >= 4:                         raw = 1
    elif m >= 2 and neutral == 0:        raw = 1
    elif m >= 1 and aero >= 1:           raw = 1
    else:                                raw = 0

    if qt03 and raw > 0:
        raw = max(0, raw - 1)
    return raw


# ── Label 150 bai ────────────────────────────────────────────────────────────
all_150 = (
    [{"idx": i+1,   "block": "RANDOM",         "art": a} for i, a in enumerate(sampled_100)] +
    [{"idx": 101+i, "block": "BALANCED_BOOST",  "art": a} for i, a in enumerate(block2_new)]
)

results = []
for row in all_150:
    art = row["art"]
    lbl = classify_v2(art.get("title","") or "", art.get("content","") or "")
    results.append({**row, "label": lbl})

dist = {0: sum(1 for r in results if r["label"]==0),
        1: sum(1 for r in results if r["label"]==1),
        2: sum(1 for r in results if r["label"]==2)}
print(f"\nDistribution v4: {dist}")
print(f"  Nhan 0: {dist[0]/150*100:.1f}%  Nhan 1: {dist[1]/150*100:.1f}%  Nhan 2: {dist[2]/150*100:.1f}%")

# ══════════════════════════════════════════════════════════════════════════════
# UPDATE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
LABEL_BG = {0: "E2EFDA", 1: "FFF2CC", 2: "FCE4D6"}
LABEL_FG = {0: "375623", 1: "7F6000", 2: "843C0C"}

def fill_c(h): return PatternFill("solid", fgColor=h)
def thin_b():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

print(f"\nUpdating Excel...")
SRC = Path(__file__).parent / "P0-03_pilot_labeling_150.xlsx"
import shutil; shutil.copy(SRC, XL)
wb = openpyxl.load_workbook(XL)
ws = wb["📋 Gán Nhãn"]

for r in results:
    idx = r["idx"]
    art = r["art"]
    lbl = r["label"]
    blk = r["block"]
    excel_row = idx + 2 if idx <= 100 else idx + 3

    a_cell = ws.cell(row=excel_row, column=1)
    if a_cell.__class__.__name__ == "MergedCell":
        continue

    base_bg = ("EBF3FB" if idx % 2 == 1 else "FFFFFF") if blk == "RANDOM" \
              else ("FFF9EC" if idx % 2 == 1 else "FFFDF5")

    # Cols A-F: content
    row_vals = {
        1: idx, 2: f"ART-{idx:03d}",
        3: art.get("publish_date",""), 4: art.get("source",""),
        5: art.get("title",""), 6: art.get("content","") or "",
    }
    for col_i, val in row_vals.items():
        c = ws.cell(row=excel_row, column=col_i)
        if c.__class__.__name__ == "MergedCell": continue
        c.value = val
        c.fill  = fill_c(base_bg)
        c.font  = Font(size=9, name="Calibri")
        c.alignment = Alignment(
            horizontal="center" if col_i <= 4 else "left",
            vertical="top", wrap_text=True)
        c.border = thin_b()

    # Col G: Claude label
    g = ws.cell(row=excel_row, column=7)
    if g.__class__.__name__ == "MergedCell": continue
    g.value     = lbl
    g.font      = Font(bold=True, color=LABEL_FG[lbl], size=11, name="Calibri")
    g.fill      = fill_c(LABEL_BG[lbl])
    g.alignment = Alignment(horizontal="center", vertical="center")
    g.border    = thin_b()

    # Cols H-J: clear
    for col_i in [8, 9, 10]:
        c = ws.cell(row=excel_row, column=col_i)
        if c.__class__.__name__ == "MergedCell": continue
        c.value = None; c.fill = fill_c("FFF2CC")
        c.font = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_b()

# Separator row 103
sep_row = 103
try: ws.unmerge_cells(f"A{sep_row}:K{sep_row}")
except: pass
ws.merge_cells(f"A{sep_row}:K{sep_row}")
c = ws[f"A{sep_row}"]
c.value = ("--- BLOCK 2: 25 bai HIGH_RISK + 25 bai MEDIUM_RISK (title-filtered) ---  "
           "Tat ca van gan nhan doc lap theo tieu chi LABELING_GUIDE_v2.md")
c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
c.fill      = fill_c("2E4057")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[sep_row].height = 22

# Footer row 154
footer_row = 154
try: ws.unmerge_cells(f"A{footer_row}:K{footer_row}")
except: pass
ws.merge_cells(f"A{footer_row}:K{footer_row}")
c = ws[f"A{footer_row}"]
c.value = (f"TONG: 150 BAI  |  Block 1 (100 random)  |  "
           f"Block 2 (25 HIGH + 25 MEDIUM title-filtered)  |  "
           f"Claude v2: 0={dist[0]}  1={dist[1]}  2={dist[2]}")
c.font      = Font(bold=True, color="1F3864", size=9, name="Calibri")
c.fill      = fill_c("D9E1F2")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[footer_row].height = 18

wb.save(XL)
print(f"Saved: {XL}")
print(f"Claude v2 final: 0={dist[0]}  1={dist[1]}  2={dist[2]}")
