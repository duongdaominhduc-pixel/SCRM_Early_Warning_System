"""
P0-03 — Rebuild Block 2: thay 50 bai cu bang 50 bai HIGH_RISK ro rang
Filter chi tren TITLE (khong can doc content) de dam bao
4 rater deu se gan Nhan 2 khong can suy nghi nhieu.

Tieu chi loc TITLE (theo guide v2):
  T1 - Bankruptcy/collapse:  bankrupt, chapter 11, insolvent, liquidat, default on debt
  T2 - Strike active:        strike begins/underway/enters day/launched, walkout begins, work stoppage
  T3 - Port/plant shutdown:  port shut/closed/closure, terminal closed, factory shut
  T4 - Vessel/cargo seized:  ship seized, vessel detained, cargo blocked, Suez blocked
  T5 - Houthi/Red Sea:       Houthi attack/fire/hit, ship attacked, vessel hijacked
  T6 - Lockdown active:      lockdown imposed/begins, city sealed, factory lockdown
  T7 - Disaster confirmed:   hit by hurricane/typhoon/flood, earthquake halts, flood shuts
  T8 - Confirmed shortage:   stockout, out of stock, acute shortage, supply crisis confirmed
"""

import json, random, re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).parent.parent
NEWS = BASE / "0. news data" / "news_2022_2024_clean_2.json"
XL   = Path(__file__).parent / "P0-03_pilot_labeling_150.xlsx"

SEED   = 42
STRATA = {"2022": 15, "2023": 38, "2024": 47}

# ── Rebuild Block 1 (same seed) ───────────────────────────────────────────────
random.seed(SEED)
with open(NEWS, "r", encoding="utf-8") as f:
    all_articles = json.load(f)

by_year = defaultdict(list)
for art in all_articles:
    yr = str(art.get("publish_date", ""))[:4]
    if yr in STRATA:
        by_year[yr].append(art)

sampled_100 = []
for yr, n in STRATA.items():
    sampled_100.extend(random.sample(by_year[yr], n))
random.shuffle(sampled_100)
block1_urls = {a["url"] for a in sampled_100}

print(f"Block 1 giữ nguyên: {len(sampled_100)} bài")

# ══════════════════════════════════════════════════════════════════════════════
# TITLE-LEVEL HIGH_RISK FILTERS (guide v2 — chỉ match trên TITLE)
# ══════════════════════════════════════════════════════════════════════════════

TITLE_HR_PATTERNS = [
    # T1 — Bankruptcy / collapse
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\bfil\w+\s+for\s+(bankruptcy|chapter\s+11)\b',
    r'\b(goes?|went|filed)\s+(bankrupt|into\s+administration|insolvent)\b',
    r'\bdefault\w*\s+on\s+(debt|payment|bond|loan)\b',
    r'\bceas\w+\s+(operations?|trading|business)\b',
    r'\bdebt\s+default\b',
    r'\bcollaps\w+\s+(of|in)\b.{0,40}\b(supply|chain|carrier|airline|freight)\b',

    # T2 — Strike ACTIVE (da bat dau, dang dien ra)
    r'\bstrike\s+(begins|started|underway|launched|enters\s+day\s*\d|enters\s+\w+\s+day|in\s+its?\s+\d+\w*\s+day)\b',
    r'\b(workers?|dockworkers?|longshoremen|port\s+workers?|rail\s+workers?|truck\s+drivers?)\s+(on\s+strike|walk\w+\s+out|begin\s+strik\w+)\b',
    r'\bwalkout\s+(begins|started|underway|hits?|halts?)\b',
    r'\bwork\s+stoppage\s+(begins|at|hits?|halts?)\b',
    r'\bpicket\s+(line|lines)\s+(block\w*|halt\w*|prevent\w*)\b',
    r'\bport\s+\w+\s+shut\w*\s+due\s+to\s+(strike|labor|walkout)\b',
    r'\bstrike\s+(halt\w*|shut\w*|crippl\w*|paralyз\w*|idl\w+)\s+\w*\s*(port|terminal|plant|factory|rail|freight)\b',
    r'\b(rail|railway|railroad)\s+strike\s+(begins|underway|halts?|shuts?)\b',

    # T3 — Port / terminal / plant SHUTDOWN
    r'\bport\s+of\s+\w[\w\s]{0,20}(shut\w*|clos\w*|halt\w*|suspend\w*)\b',
    r'\b(port|terminal|harbor)\s+(shut\s*down|closure|closed|halted|suspended|blocked)\b',
    r'\b(shut\s*down|closure|closed|halted)\s+(port|terminal|harbor|plant|factory)\b',
    r'\b(plant|factory|facilit\w+|assembly)\s+(shut\w*|clos\w*|halt\w*|suspend\w*|idl\w+)\b',
    r'\bproduction\s+(halted|suspended|shut\s+down|stopped)\b',
    r'\bport\s+(operations?|activity)\s+(halt\w*|suspend\w*|ceas\w*)\b',

    # T4 — Vessel / cargo seized / blocked
    r'\b(ship|vessel|cargo|tanker|freighter)\s+(seized|detained|impounded|confiscated|blocked|stranded)\b',
    r'\bever\s*given\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*|jam\w*)\b',
    r'\b\d+\s+(ships?|vessels?)\s+(strand\w*|block\w*|wait\w*)\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?)\s+(strand\w*|block\w*|anchor\w*|wait\w*)\b',

    # T5 — Houthi / Red Sea ACTIVE attack / reroute
    r'\bhouthi\w*\s+(attack\w*|fire\w*|hit\s+|struck|seized|hijack\w*|target\w*)\b',
    r'\b(ship|vessel|tanker)\s+(attack\w*|hit\s+by|struck\s+by|seized\s+by|hijack\w*)\b.{0,60}\b(houthi|red\s+sea|yemen)\b',
    r'\bcrew\s+(evacuated|rescued|abandon\w*)\b.{0,60}\b(red\s+sea|houthi|attack)\b',
    r'\b(maersk|hapag|msc|cma\s*cgm|evergreen|cosco|zim)\s+\w{0,15}\s*(suspend\w*|halt\w*|divert\w*|reroute\w*|avoid\w*)\b',
    r'\bshipping\s+(route\w*|line\w*|service\w*)\s+(suspend\w*|halt\w*|divert\w*|reroute\w*)\b',

    # T6 — Lockdown ACTIVE
    r'\b(covid|covid-19|coronavirus)\s+(lockdown|shutdown|closure)\b',
    r'\blockdown\s+(hits?|halt\w*|shut\w*|crippl\w*|jeopardiz\w*|disrupt\w*)\b',
    r'\b(city|cities|region|province|district)\s+(seal\w*|lock\w+\s+down|shut\w+\s+down)\b',
    r'\bzero[- ]covid\s+(lockdown|shutdown|policy\s+halt\w*)\b',

    # T7 — Natural disaster CONFIRMED impact
    r'\b(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami|cyclone)\s+\w{0,20}\s*(halt\w*|shut\w*|clos\w*|disrupt\w*|damage\w*|destroy\w*|crippl\w*)\b',
    r'\b(halt\w*|shut\w*|clos\w*|disrupt\w*|damage\w*|destroy\w*)\s+\w{0,20}\s*(hurricane|typhoon|flood|earthquake|wildfire)\b',
    r'\b(flood|earthquake|fire)\s+(shut\w*|halt\w*|clos\w*)\s+(port|factory|plant|facilit\w+|road|bridge)\b',
    r'\b(port|factory|plant|highway|road|bridge)\s+(shut\w*|halt\w*|clos\w*)\s+\w{0,20}\s*(flood|earthquake|fire|storm)\b',

    # T8 — Confirmed shortage / stockout
    r'\b(stockout|out\s+of\s+stock|ran\s+out\s+of)\b',
    r'\b(critical|acute|severe)\s+shortage\b',
    r'\bsupply\s+(crisis|collapse|failure)\s+(hit\w*|affect\w*|crippl\w*)\b',
    r'\b(chip|semiconductor|fuel|energy|gas)\s+shortage\s+(crippl\w*|halt\w*|shut\w*)\b',
]


def title_hr_score(title: str) -> int:
    """Dem so HIGH_RISK pattern match tren title."""
    t = title.lower()
    return sum(1 for p in TITLE_HR_PATTERNS if re.search(p, t, re.IGNORECASE))


# ── Filter corpus: lay bai NGOAI Block 1, co title HR score >= 1 ──────────────
pool = [a for a in all_articles
        if str(a.get("publish_date", ""))[:4] in STRATA
        and a["url"] not in block1_urls]

hr_pool = []
for art in pool:
    sc = title_hr_score(art.get("title", "") or "")
    if sc >= 1:
        hr_pool.append((sc, art))

# Sort giam dan theo score (bai ro rang nhat dau)
hr_pool.sort(key=lambda x: -x[0])

print(f"\nPool bai co title HR signal >= 1: {len(hr_pool)} bai")
print(f"  Score >= 2: {sum(1 for sc,_ in hr_pool if sc>=2)}")
print(f"  Score >= 3: {sum(1 for sc,_ in hr_pool if sc>=3)}")

# Preview top 20
print("\nTop 20 bai HR ro nhat:")
for i, (sc, art) in enumerate(hr_pool[:20], 1):
    print(f"  [{i:02d}] sc={sc} | {art.get('title','')[:90]}")

# Sample 50 tu top 150 bai ro rang nhat (khong chi lay top 50 tuyet doi de tranh bias)
top_pool_size = min(150, len(hr_pool))
top_hr = [art for _, art in hr_pool[:top_pool_size]]

random.seed(SEED + 2)
if len(top_hr) >= 50:
    block2_new = random.sample(top_hr, 50)
else:
    block2_new = top_hr
    print(f"WARNING: chi co {len(top_hr)} bai HR du tieu chuan, lay tat ca")

print(f"\nBlock 2 moi: {len(block2_new)} bai (title HR score >= 1)")
score_dist = {1:0, 2:0, 3:0, 4:0}
for art in block2_new:
    sc = min(title_hr_score(art.get("title","")), 4)
    score_dist[sc] = score_dist.get(sc, 0) + 1
print(f"  Phan phoi score: {score_dist}")

# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFIER v2 (copy tu relabel_claude_v2.py)
# ══════════════════════════════════════════════════════════════════════════════
STRIKE_ACTIVE = [
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day\s*\d|enters\s+\w+\s+day|launched|action\s+begins)\b',
    r'\b(workers?|dockworkers?|longshoremen)\s+(are\s+)?(on\s+strike|walk\w+\s+out|went\s+on\s+strike|began\s+strik\w+)\b',
    r'\bwalkout\s+(begins|started|underway)\b',
    r'\bwork\s+stoppage\s+(begins|started|underway|in\s+effect)\b',
    r'\bpicket\s+line\b',
    r'\bstrikers?\s+(block|prevent|halt|shut)\b',
]
RED_SEA_ACTIVE = [
    r'\bhouthi\w*\s+(attack\w*|fire\w*|hit\s+|struck|seized|hijack\w*)\b',
    r'\b(maersk|hapag|msc|cma\s+cgm|evergreen|cosco|zim)\s+\w*\s*(suspend\w*|halt\w*|divert\w*|reroute\w*)\b',
    r'\bships?\s+(divert\w*|reroute\w*)\b.{0,60}\b(cape|good\s+hope|africa)\b',
    r'\bcrew\s+(evacuated|rescued)\b.{0,60}\b(red\s+sea|houthi)\b',
]
PORT_CLOSED = [
    r'\b(port|terminal|harbor)\s+(is\s+)?(clos\w+|shut\s*down|halt\w*|suspend\w*)\b',
    r'\bport\s+of\s+\w+\s+(clos\w+|shut\w*|halt\w*)\b',
]
PLANT_CLOSED = [
    r'\b(plant|factory|facilit\w+)\s+(shut\w*|clos\w*|halt\w*|idl\w+)\b',
    r'\bproduction\s+(halt\w*|suspend\w*|shut\w*)\b',
]
BANKRUPTCY_ACTIVE = [
    r'\bfil\w+\s+for\s+(bankruptcy|chapter\s+11)\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\bceas\w+\s+(operat\w*|trading|business)\b',
    r'\bdefault\w*\s+on\s+(debt\w*|payment\w*)\b',
]
DISASTER_ACTIVE = [
    r'\b(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami)\s+\w*\s*(hit\w*|struck|devastat\w*|halt\w*|shut\w*)\b',
    r'\b(hit|struck|devastat\w*)\s+by\s+(hurricane|typhoon|flood\w*|earthquake)\b',
]
SANCTION_ACTIVE = [
    r'\b(sanction\w*|embargo|ban)\s+(in\s+effect|effective|enforced|imposed)\b',
    r'\b(us|eu|un|uk)\s+imposes?\s+(sanction\w*|ban|embargo)\b',
]
LOCKDOWN_ACTIVE = [
    r'\b(lockdown|lock\s+down)\s+(in\s+effect|imposed|begins|started|underway)\b',
    r'\bcovid\s*-?\s*19?\s+(lockdown|shutdown|closure)\b',
]
SHORTAGE_CONFIRMED = [
    r'\b(stockout|out\s+of\s+stock|ran\s+out\s+of|depleted)\b',
    r'\b(critical|acute|severe)\s+shortage\b',
    r'\bsupply\s+gap\s+(widen\w*|grow\w*)\b',
]
CARGO_SEIZED = [
    r'\b(cargo|vessel|ship)\s+(seized|detained|impounded|blocked)\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*)\b',
]
SEVERE_DELAY = [
    r'\bdelays?\s+of\s+(several\s+)?(weeks?|months?)\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?)\s+(wait\w*|strand\w*)\b',
]
ALL_HIGH = (STRIKE_ACTIVE + RED_SEA_ACTIVE + PORT_CLOSED + PLANT_CLOSED +
            BANKRUPTCY_ACTIVE + DISASTER_ACTIVE + SANCTION_ACTIVE +
            LOCKDOWN_ACTIVE + SHORTAGE_CONFIRMED + CARGO_SEIZED + SEVERE_DELAY)

MEDIUM_PATTERNS = [
    r'\b(contract\s+talks?|labor\s+negotiat\w*|union\s+negotiat\w*)\b',
    r'\bstrike\s+(threat|vote|authoriz\w*|warning|loom\w*|possible|risk)\b',
    r'\b(workers?|union)\s+(vote|threaten|warn\w*)\s+\w*\s*(strike|walkout)\b',
    r'\bport\s+congestion\b',
    r'\b(increas\w+|ris\w+|worsening)\s+(congestion|delay\w*|backlog)\b',
    r'\bshipping\s+(delay\w*|backlog|bottleneck|disruption)\b',
    r'\b(risk|threat|fear\w*)\s+(of\s+)?(shortage|shortfall|disruption)\b',
    r'\bsupplier\s+(issue\w*|problem\w*|strain\w*|distress\w*)\b',
    r'\btrade\s+(war|dispute|tension\w*)\b',
    r'\btariff\s+(hike|increase|propos\w*|threat\w*)\b',
    r'\b(storm|typhoon|hurricane)\s+(threaten\w*|approach\w*|warn\w*|watch)\b',
    r'\b(geopolitical)\s+(tension\w*|uncertainty|instability)\b',
    r'\b(financial\s+distress|cash\s+crunch)\b',
]
NEUTRAL_HARD = [
    r'\bartificial\s+intelligence\b', r'\bmachine\s+learning\b',
    r'\bdigital\s+(transform\w*|solution\w*|platform\w*)\b',
    r'\bwebinar\b', r'\bpodcast\b', r'\bwhitepaper\b',
    r'\baward\w*\s+(winner|recipient)\b',
    r'\bnew\s+(ceo|cfo|coo|vp|director)\b',
    r'\bmarket\s+(share|size|growth|forecast)\b',
    r'\bprofit\s+(rise\w*|increas\w*|beat\w*)\b',
    r'\bipo\b', r'\bfunding\s+round\b',
]
AEROSPACE_RELATED = [
    r'\baerospace\b', r'\baviation\b', r'\baircraft\b', r'\bairline\b',
    r'\bairport\b', r'\bair\s+cargo\b', r'\bair\s+freight\b',
    r'\bboeing\b', r'\bairbus\b', r'\bdefense\b',
    r'\bport\b', r'\bshipping\b', r'\bfreight\b', r'\bcontainer\b',
    r'\bsuez\b', r'\bpanama\b', r'\bred\s+sea\b',
    r'\btranspacific\b', r'\btransatlantic\b',
    r'\bsemiconductor\b', r'\benergy\s+(supply|crisis)\b',
    r'\bsteel\b', r'\bcopper\b', r'\brare\s+earth\b',
]
NON_AEROSPACE_HARD = [
    r'\bautomotive\b', r'\bauto\s+(industry|maker|plant)\b',
    r'\belectric\s+vehicle\b', r'\bev\s+(battery|maker)\b',
    r'\bfarm\w*\b', r'\bagricultur\w*\b', r'\bgrain\b', r'\bwheat\b',
    r'\bpharmaceutical\b', r'\bdrug\s+(supply|shortage)\b',
    r'\bfood\s+(supply|chain|shortage)\b',
    r'\bretail\b', r'\becommerce\b', r'\bfashion\b',
]
FUTURE_HEDGE = [
    r'\b(could|may|might|would|expect\w*|forecast\w*|anticipat\w*)\b',
    r'\b(risk\s+of|threat\s+of|fear\s+of)\b',
    r'\blooming\b', r'\bpotential\b', r'\bpossible\b', r'\bpending\b',
]


def cnt(text, patterns):
    t = text.lower()
    return sum(1 for p in patterns if re.search(p, t, re.IGNORECASE))


def classify_v2(title, content):
    full = f"{title} {title} {title} {content}"
    tl   = title.lower()
    neutral = cnt(full, NEUTRAL_HARD)
    if neutral >= 2:
        return 0
    aero    = cnt(full, AEROSPACE_RELATED)
    non_aero= cnt(full, NON_AEROSPACE_HARD)
    qt03    = (non_aero >= 1 and aero == 0)

    h_title = cnt(tl, ALL_HIGH)
    h_full  = cnt(full, ALL_HIGH)
    h_score = h_title * 3 + h_full
    m_title = cnt(tl, MEDIUM_PATTERNS)
    m_full  = cnt(full, MEDIUM_PATTERNS)
    m_score = m_title * 2 + m_full
    hedge   = cnt(tl + " " + content[:500], FUTURE_HEDGE)

    if h_score >= 5 or h_title >= 2:             raw = 2
    elif h_score >= 3 and hedge <= 2:            raw = 2
    elif h_score >= 2 and m_score >= 2 and hedge <= 1: raw = 2
    elif h_score >= 2 or (h_score==1 and m_score>=4): raw = 1
    elif m_score >= 4:                            raw = 1
    elif m_score >= 2 and neutral == 0:           raw = 1
    elif m_score >= 1 and aero >= 1:              raw = 1
    else:                                         raw = 0

    if qt03 and raw > 0:
        raw = max(0, raw - 1)
    return raw


# ── Label tat ca 150 bai moi ─────────────────────────────────────────────────
all_150_new = (
    [{"idx": i+1,    "block": "RANDOM",           "art": a} for i, a in enumerate(sampled_100)] +
    [{"idx": 101+i,  "block": "HIGH_RISK_TITLE",  "art": a} for i, a in enumerate(block2_new)]
)

print(f"\nLabeling 150 articles v3 (Block 2 = title-filtered HR)...")
results = []
for row in all_150_new:
    art  = row["art"]
    title = art.get("title","") or ""
    lbl  = classify_v2(title, art.get("content","") or "")
    results.append({"idx": row["idx"], "block": row["block"],
                    "label": lbl, "art": art})

dist = {0: sum(1 for r in results if r["label"]==0),
        1: sum(1 for r in results if r["label"]==1),
        2: sum(1 for r in results if r["label"]==2)}
print(f"Distribution v3: {dist}")
print(f"Nhan 2 rate: {dist[2]/150*100:.1f}%")

# Preview Block 2
print("\nBlock 2 preview (50 bai title-filtered):")
for r in results[100:]:
    print(f"  [{r['idx']:03d}] L={r['label']} | {r['art'].get('title','')[:85]}")

# ══════════════════════════════════════════════════════════════════════════════
# UPDATE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def fill_c(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_b():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

LABEL_BG = {0: "E2EFDA", 1: "FFF2CC", 2: "FCE4D6"}
LABEL_FG = {0: "375623", 1: "7F6000", 2: "843C0C"}
BLOCK2_ROW_ODD  = "FFF9EC"
BLOCK2_ROW_EVEN = "FFFDF5"

print(f"\nUpdating Excel: {XL}")
wb = openpyxl.load_workbook(XL)
ws = wb["📋 Gán Nhãn"]

# Map idx -> result
result_map = {r["idx"]: r for r in results}

# ── Pass 1: update content (A-F) va nhan (G) cho 150 rows ────────────────────
excel_data_row = 3
for r in results:
    idx   = r["idx"]
    art   = r["art"]
    lbl   = r["label"]
    blk   = r["block"]

    # Tinh row trong Excel
    # Block 1: row 3..102, Block 2: row 104..153 (row 103 = separator)
    if idx <= 100:
        excel_row = idx + 2
    else:
        excel_row = idx + 3   # +3 vi co separator row 103

    # Check merged cell
    a_cell = ws.cell(row=excel_row, column=1)
    if a_cell.__class__.__name__ == "MergedCell":
        continue

    base_bg = (BLOCK2_ROW_ODD if idx % 2 == 1 else BLOCK2_ROW_EVEN) if blk == "HIGH_RISK_TITLE" \
              else ("EBF3FB" if idx % 2 == 1 else "FFFFFF")

    # Update A-F (content columns)
    content = art.get("content","") or ""
    row_vals = {
        1: idx,
        2: f"ART-{idx:03d}",
        3: art.get("publish_date",""),
        4: art.get("source",""),
        5: art.get("title",""),
        6: content,
    }
    for col_idx, val in row_vals.items():
        c = ws.cell(row=excel_row, column=col_idx)
        if c.__class__.__name__ == "MergedCell":
            continue
        c.value = val
        c.fill  = fill_c(base_bg)
        c.font  = Font(size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center" if col_idx <= 4 else "left",
                                vertical="top", wrap_text=True)
        c.border = thin_b()

    # Col G: nhan Claude v2
    g = ws.cell(row=excel_row, column=7)
    if g.__class__.__name__ == "MergedCell":
        continue
    g.value     = lbl
    g.font      = Font(bold=True, color=LABEL_FG[lbl], size=11, name="Calibri")
    g.fill      = fill_c(LABEL_BG[lbl])
    g.alignment = Alignment(horizontal="center", vertical="center")
    g.border    = thin_b()

    # Col H, I, J: xoa nhan cu
    for col_idx in [8, 9, 10]:
        c = ws.cell(row=excel_row, column=col_idx)
        if c.__class__.__name__ == "MergedCell":
            continue
        c.value     = None
        c.fill      = fill_c("FFF2CC")
        c.font      = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_b()

# ── Update separator row (103) ────────────────────────────────────────────────
# Unmerge truoc neu can
sep_row = 103
try:
    ws.unmerge_cells(f"A{sep_row}:K{sep_row}")
except:
    pass
ws.merge_cells(f"A{sep_row}:K{sep_row}")
c = ws[f"A{sep_row}"]
c.value = ("--- 50 BAI BLOCK 2 (TITLE-FILTERED HIGH_RISK) ---  "
           "Cac bai nay duoc loc theo TITLE co signal HIGH_RISK ro rang theo guide v2. "
           "Tat ca 4 raters van gan nhan doc lap theo tieu chi.")
c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
c.fill      = fill_c("843C0C")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[sep_row].height = 22

# ── Update footer ─────────────────────────────────────────────────────────────
footer_row = 154
try:
    ws.unmerge_cells(f"A{footer_row}:K{footer_row}")
except:
    pass
ws.merge_cells(f"A{footer_row}:K{footer_row}")
c = ws[f"A{footer_row}"]
c.value = (f"TONG: 150 BAI  |  Block 1 (Random Stratified): 100 bai  "
           f"|  Block 2 (Title-Filtered HIGH_RISK): 50 bai  "
           f"|  Claude v2 dist: 0={dist[0]}  1={dist[1]}  2={dist[2]}")
c.font      = Font(bold=True, color="1F3864", size=9, name="Calibri")
c.fill      = fill_c("D9E1F2")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[footer_row].height = 18

wb.save(XL)
print(f"Saved: {XL}")
print(f"\nTong ket:")
print(f"  Block 1 (random): 100 bai | Block 2 (title HR): 50 bai")
print(f"  Claude v2 labels: 0={dist[0]}  1={dist[1]}  2={dist[2]}")
print(f"  Nhan 2 = {dist[2]} bai ({dist[2]/150*100:.1f}%)")
