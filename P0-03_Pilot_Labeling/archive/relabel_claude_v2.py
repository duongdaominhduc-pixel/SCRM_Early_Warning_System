"""
P0-03 — Xoa nhan cu, gan lai cot Claude theo LABELING_GUIDE_v2.md

4 quy tac moi:
  QT-01: Dinh cong bat dau (ngay 1) -> Nhan 2 ngay
  QT-02: Tau bi tan cong / hang tau doi tuyen Bien Do -> Nhan 2
  QT-03: Khong lien quan aerospace -> giam 1 bac nhan
  QT-04: Ngon ngu manh (crisis/severe) nhung su kien chua xay ra -> Nhan 1 (khong phai 2)
"""

import json, random, re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = Path(__file__).parent.parent
NEWS = BASE / "0. news data" / "news_2022_2024_clean_2.json"
XL   = Path(__file__).parent / "P0-03_pilot_labeling_150.xlsx"

SEED   = 42
STRATA = {"2022": 15, "2023": 38, "2024": 47}

random.seed(SEED)

# ── Rebuild same 150 articles (same seeds) ────────────────────────────────────
with open(NEWS, "r", encoding="utf-8") as f:
    all_articles = json.load(f)

by_year = defaultdict(list)
for art in all_articles:
    yr = str(art.get("publish_date", ""))[:4]
    if yr in STRATA:
        by_year[yr].append(art)

sampled_100 = []
for yr, n in STRATA.items():
    sampled_100.extend(random.sample(by_year[yr], n))
random.shuffle(sampled_100)
sampled_urls = {a["url"] for a in sampled_100}

# Block 2 HIGH_RISK oversampling (seed+1)
HIGH_RISK_PATTERNS_SCORE = [
    r'\b(port|ports|terminal)\s+(shut\s*down|clos\w+|halt\w*|block\w*|suspend\w*)\b',
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day|hit\s+|crippl\w+)\b',
    r'\b(workers?|dockworkers?|longshoremen)\s+(on\s+strike|walk\w*\s+out|halted\s+work)\b',
    r'\bwork\s+stoppage\b', r'\bwalkout\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\b(disrupted|halted|suspended|blocked|stranded|grounded|paralyzed|crippled)\b',
    r'\b(flood|earthquake|typhoon|hurricane|wildfire|tsunami)\s+\w*\s*(hit|struck|devastat\w*|halt\w*)\b',
    r'\b(sanction\w*|embargo|ban)\s+(imposed|effective|enforced)\b',
    r'\b(plant|factory|facilit\w+)\s+(shut\w*|clos\w*|halt\w*)\b',
    r'\b(severe|critical|acute)\s+(shortage|shortfall)\b',
    r'\b(vessel|ship|container)\s+(strand\w*|block\w*|detain\w*|seiz\w*)\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?)\s+(wait\w*|strand\w*)\b',
]
def hr_score(art):
    txt = f"{art.get('title','')*3} {art.get('content','')}"
    return sum(1 for p in HIGH_RISK_PATTERNS_SCORE if re.search(p, txt, re.I))

pool_rest = [a for a in all_articles
             if str(a.get("publish_date",""))[:4] in STRATA
             and a["url"] not in sampled_urls]
scored = sorted([(hr_score(a), a) for a in pool_rest], key=lambda x: -x[0])
top200 = [a for sc, a in scored if sc >= 3][:200]
random.seed(SEED + 1)
over50 = random.sample(top200, 50) if len(top200) >= 50 else [a for _, a in scored[:50]]

all_150 = (
    [{"idx": i+1,    "block": "RANDOM",          "art": a} for i, a in enumerate(sampled_100)] +
    [{"idx": 101+i,  "block": "HIGH_RISK_BOOST",  "art": a} for i, a in enumerate(over50)]
)

# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFIER v2 — theo LABELING_GUIDE_v2.md
# ══════════════════════════════════════════════════════════════════════════════

# --- [A] Su kien DA XAY RA -> Nhan 2 (truoc khi ap QT-03) -------------------

# QT-01: Dinh cong DA bat dau
STRIKE_ACTIVE = [
    r'\bstrike\s+(begins|started|underway|continues|enters\s+day\s*\d|launched|action\s+begins)\b',
    r'\b(workers?|dockworkers?|longshoremen|truckers?|employees?)\s+(are\s+)?(on\s+strike|walk\w+\s+out|went\s+on\s+strike|began\s+strik\w+)\b',
    r'\bwalkout\s+(begins|started|underway)\b',
    r'\bwork\s+stoppage\s+(begins|started|underway|in\s+effect)\b',
    r'\bpicket\s+line\b',
    r'\bstrikers?\s+(block|prevent|halt|shut)\b',
    r'\bstrike\s+(halt\w*|shut\w*|crippl\w*|paralyз\w*|disrupt\w*)\b',
    r'\bdinh\s+cong\s+(bat\s+dau|dang\s+xay\s+ra)\b',
]
# QT-02: Bien Do / Houthi - tau bi tan cong HOAC hang tau DA doi tuyen
RED_SEA_ACTIVE = [
    r'\b(houthi\w*)\s+(attack\w*|fire\w*|hit\s+|struck|seized|hijack\w*)\b',
    r'\b(attack\w*|fire\w*|struck|seized|hijack\w*)\b.{0,60}\b(red\s+sea|gulf\s+of\s+aden)\b',
    r'\b(maersk|hapag|msc|cma\s+cgm|evergreen|cosco|zim|yang\s+ming)\s+\w*\s*(suspend\w*|halt\w*|divert\w*|reroute\w*|avoid\w*|skip\w*)\b',
    r'\b(carrier\w*|shipping\s+line\w*|ship\w*)\s+\w*\s*(suspend\w*|halt\w*|divert\w*|reroute\w*)\b.{0,80}\b(red\s+sea|suez)\b',
    r'\bships?\s+(divert\w*|reroute\w*)\b.{0,60}\b(cape|good\s+hope|africa)\b',
    r'\bvoyage\s+around\s+(africa|cape)\b',
]
# Cang / nha may da dong cua
PORT_CLOSED = [
    r'\b(port|terminal|harbor)\s+(is\s+)?(clos\w+|shut\s*down|halt\w*|suspend\w*|block\w*)\b',
    r'\b(clos\w+|shut\s*down|halt\w*|suspend\w*)\s+(port|terminal|operations?)\b',
    r'\bport\s+of\s+\w+\s+(clos\w+|shut\w*|halt\w*)\b',
]
PLANT_CLOSED = [
    r'\b(plant|factory|facilit\w+|warehouse|assembly)\s+(is\s+)?(shut\w*|clos\w*|halt\w*|idl\w+|suspend\w*)\b',
    r'\b(shut\w*|clos\w*|halt\w*|suspend\w*)\s+(plant|factory|facilit\w+|production|operations?)\b',
    r'\bproduction\s+(halt\w*|suspend\w*|shut\w*|stop\w*)\b',
]
# Pha san / vo no da cong bo
BANKRUPTCY_ACTIVE = [
    r'\bfil\w+\s+for\s+(bankruptcy|chapter\s+11|insolvenc\w*)\b',
    r'\b(bankrupt\w*|chapter\s+11|insolvenc\w*|liquidat\w*)\b',
    r'\bceas\w+\s+(operat\w*|trading|business)\b',
    r'\bwent\s+(bankrupt|into\s+administration|insolvent)\b',
    r'\bdefault\w*\s+on\s+(debt\w*|payment\w*|loan\w*|bond\w*)\b',
    r'\bdebt\s+(default|restructur\w*)\b',
]
# Thien tai da do bo / gay thiet hai
DISASTER_ACTIVE = [
    r'\b(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami|cyclone|tornado)\s+\w{0,15}\s*(hit\w*|struck|devastat\w*|destroy\w*|slam\w*|batter\w*|sweep\w*|ravag\w*)\b',
    r'\b(hit|struck|devastat\w*|destroy\w*|slam\w*)\s+by\s+(hurricane|typhoon|flood\w*|earthquake|wildfire|tsunami)\b',
    r'\b(flood\w*|earthquake|fire)\s+(halt\w*|shut\w*|clos\w*|disrupt\w*|damage\w*)\s+(port|factory|plant|facilit\w+|road|bridge|infra)\b',
]
# Lenh trung phat / cam van da hieu luc
SANCTION_ACTIVE = [
    r'\b(sanction\w*|embargo|ban|restriction\w*)\s+(now\s+)?(in\s+effect|effective|enforced|imposed|applied|implement\w*)\b',
    r'\b(us|eu|un|uk)\s+(imposes?|imposed|enacted|enacted)\s+(sanction\w*|ban|embargo)\b',
    r'\bnew\s+sanction\w*\s+on\b',
    r'\bsanction\w*\s+(target\w*|hit\w*|block\w*)\s+(russian|iranian|chinese|north\s+korean)\b',
]
# Lockdown / phong toa
LOCKDOWN_ACTIVE = [
    r'\b(lockdown|lock\s+down)\s+(in\s+effect|imposed|announced|begins|started|underway)\b',
    r'\b(city|region|province|district|area)\s+(lock\w+|shut\w+|seal\w+)\b',
    r'\bcovid\s*-?\s*19?\s+(lockdown|shutdown|closure)\b',
    r'\bzero\s*-?\s*covid\s+(policy|lockdown|measures?)\b',
    r'\bquarantine\s+(imposed|mandatory|enforced)\b',
]
# Thieu hut da xac nhan nghiem trong
SHORTAGE_CONFIRMED = [
    r'\b(stockout|out\s+of\s+stock|ran\s+out\s+of|depleted)\b',
    r'\b(critical|acute|severe)\s+shortage\b',
    r'\bsupply\s+gap\s+(widen\w*|grow\w*|confirm\w*)\b',
    r'\bshortage\s+(hit\w*|affect\w*|forc\w*|caus\w*)\b.{0,60}\b(production|assembly|manufactur)\b',
    r'\bsupply\s+(crunch|crisis)\s+(is\s+)?(here|here|real|happening|worsening)\b',
]
# Tau / hang bi giu / chan
CARGO_SEIZED = [
    r'\b(cargo|vessel|ship|container)\s+(seized|detained|impounded|confiscated|blocked)\b',
    r'\b(seized|detained|impounded)\s+(cargo|vessel|ship|container)\b',
    r'\bsuez\s+canal\s+(block\w*|clos\w*|strand\w*)\b',
    r'\bever\s*given\b',
]
# Delay da xac nhan nghiem trong (theo tuan / thang)
SEVERE_DELAY = [
    r'\bdelays?\s+of\s+(several\s+)?(weeks?|months?)\b',
    r'\b(weeks?|months?)\s+of\s+delay\w*\b',
    r'\bwait\w*\s+(weeks?|months?)\s+for\b',
    r'\b\d+\s*-?\s*(week|month)\s+delay\b',
    r'\bhundreds?\s+of\s+(ships?|vessels?|containers?)\s+(wait\w*|strand\w*|back\w*log\w*|anchor\w*|queue\w*)\b',
]

ALL_HIGH_PATTERNS = (
    STRIKE_ACTIVE + RED_SEA_ACTIVE + PORT_CLOSED + PLANT_CLOSED +
    BANKRUPTCY_ACTIVE + DISASTER_ACTIVE + SANCTION_ACTIVE +
    LOCKDOWN_ACTIVE + SHORTAGE_CONFIRMED + CARGO_SEIZED + SEVERE_DELAY
)

# --- [B] Su kien DANG HINH THANH -> Nhan 1 -----------------------------------
MEDIUM_PATTERNS = [
    # Dinh cong: dam phan / bieu quyet / canh bao
    r'\b(contract\s+talks?|labor\s+negotiat\w*|union\s+negotiat\w*|collective\s+bargaining)\b',
    r'\bstrike\s+(threat|vote|authoriz\w*|warning|loom\w*|possible|risk|could|may)\b',
    r'\b(workers?|union)\s+(vote|threaten|warn\w*|consider\w*|prepar\w*)\s+\w{0,20}\s*(strike|walkout|action)\b',
    r'\bstrike\s+authoriz\w+\b',
    r'\bcontract\s+(expir\w*|deadline|negotiat\w*|impasse|breakdown)\b',
    # Bien Do: canh bao nhung chua tan cong / chua doi tuyen
    r'\b(warn\w*|alert\w*|caution\w*|advis\w*)\b.{0,60}\b(red\s+sea|gulf\s+of\s+aden|houthi)\b',
    r'\b(houthi\w*)\s+(threat\w*|warn\w*|target\w*|could\s+attack)\b',
    r'\binsurance\s+premium\w*\s+(rise\w*|increas\w*|surge\w*)\b.{0,60}\b(red\s+sea|gulf)\b',
    # Thue quan / trung phat dang de xuat
    r'\b(propos\w+|plan\w*|consider\w*|draft\w*|mulling|weighing)\s+(tariff|sanction\w*|restriction|ban|levy)\b',
    r'\btariff\s+(hike|increase|propos\w*|threat\w*|possible|could)\b',
    r'\btrade\s+(war|dispute|tension\w*|conflict|friction)\b',
    r'\b(us|eu|un|uk)\s+(consider\w*|plan\w*|weigh\w*|discuss\w*)\s+(sanction\w*|ban|tariff)\b',
    # Thien tai dang tien den
    r'\b(storm|typhoon|hurricane|flood|cyclone)\s+(approach\w*|threaten\w*|head\w+\s+toward|forecast\w*|warn\w*|watch|warning|expected\s+to\s+hit)\b',
    r'\b(approach\w*|heading\s+toward|threaten\w*|forecast\w*|expected\s+to\s+hit)\b.{0,80}\b(port|coast|region|supply\s+chain)\b',
    # Tac nghen cang tang
    r'\bport\s+congestion\b',
    r'\b(increas\w+|grow\w+|ris\w+|worsening|escalat\w+)\s+(congestion|delay\w*|backlog)\b',
    r'\bshipping\s+(delay\w*|backlog|bottleneck|disruption)\b',
    # Thieu hut dang hinh thanh
    r'\b(risk|threat|fear\w*|concern\w*)\s+(of\s+)?(shortage|shortfall|disruption|delay)\b',
    r'\bshortage\s+(risk|concern\w*|fear\w*|loom\w*|possible|potential|ahead)\b',
    r'\b(tight|thin|low|lean)\s+(supply|inventory|stock)\b',
    r'\binventory\s+(strain|pressure|concern|crunch)\b',
    # Nha cung cap gap kho khan
    r'\bsupplier\s+(issue\w*|problem\w*|concern\w*|strain\w*|stress\w*|challeng\w*|struggl\w*|distress\w*)\b',
    r'\b(financial\s+distress|cash\s+crunch|liquidity\s+crisis)\b',
    # Lead time tang
    r'\b(longer|extended|stretched|growing)\s+lead\s+time\w*\b',
    r'\blead\s+time\w*\s+(increase\w*|grow\w*|worsen\w*|stretch\w*)\b',
    # Chinh sach thuong mai dang leo thang
    r'\b(geopolitical|trade)\s+(tension\w*|uncertainty|instability|dispute\w*)\b',
    r'\bsupply\s+chain\s+(risk\w*|vulnerab\w*|concern\w*|warn\w*)\b',
    # Panama / kenh dao thap nuoc (chua dong nhung bi han che)
    r'\bpanama\s+canal\s+(restrict\w*|drought\w*|low\w*\s+water|limit\w*|backlog)\b',
]

# --- [C] Strongly neutral -> Nhan 0 ------------------------------------------
NEUTRAL_HARD = [
    r'\bartificial\s+intelligence\b', r'\bmachine\s+learning\b',
    r'\bdigital\s+(transform\w*|solution\w*|platform\w*)\b',
    r'\bblockchain\b', r'\binternet\s+of\s+things\b',
    r'\bwebinar\b', r'\bpodcast\b', r'\bwhitepaper\b',
    r'\b\d{4}\s+annual\s+report\b',
    r'\baward\w*\s+(winner|recipient|ceremony)\b',
    r'\bnew\s+(ceo|cfo|coo|vp|president|director|officer)\b',
    r'\bappointment\b.{0,60}\b(ceo|cfo|vp|director)\b',
    r'\bjoint\s+venture\s+(formed|announced|created|finalized)\b',
    r'\bpartnership\s+(formed|announced|signed)\b.{0,60}(?!risk|disrupt)',
    r'\bmarket\s+(share|size|growth|forecast|outlook)\b',
    r'\brevenue\s+(grow\w*|increas\w*|reach\w*)\b',
    r'\bprofit\s+(rise\w*|increas\w*|beat\w*|outlook)\b',
    r'\bipo\b', r'\bfunding\s+round\b', r'\bseries\s+[abcde]\s+round\b',
]

# --- [D] Aerospace-adjacent keywords (KHONG ap QT-03 cho cac tu khoa nay) ----
AEROSPACE_RELATED = [
    r'\baerospace\b', r'\baviation\b', r'\baircraft\b', r'\bairline\b',
    r'\bairport\b', r'\bairfreight\b', r'\bair\s+cargo\b', r'\bair\s+freight\b',
    r'\bboeing\b', r'\bairbus\b', r'\bdefense\b', r'\bmilitary\b',
    r'\bsatellite\b', r'\bspacecraft\b', r'\brocket\b',
    # Nguyen lieu dung chung aerospace
    r'\btitanium\b', r'\baluminum\s+alloy\b', r'\bcomposite\s+material\b',
    # Van chuyen chung (khong gioi han nganh)
    r'\bport\b', r'\bshipping\b', r'\bfreight\b', r'\bcontainer\b',
    r'\bsuez\b', r'\bpanama\b', r'\bred\s+sea\b', r'\bblack\s+sea\b',
    r'\btranspacific\b', r'\btransatlantic\b',
    r'\bair\s+cargo\b', r'\bair\s+freight\b',
    # Nganh co ripple effect ro
    r'\bsemiconductor\b', r'\bchip\s+shortage\b',
    r'\benergy\s+(supply|crisis|shortage)\b',
    r'\bsteel\b', r'\bcopper\b', r'\brare\s+earth\b',
]
# Nganh khong lien quan -> ap QT-03 giam 1 bac
NON_AEROSPACE_HARD = [
    r'\bautomotive\b', r'\bauto\s+(industry|maker|plant)\b',
    r'\belectric\s+vehicle\b', r'\bev\s+(battery|maker)\b',
    r'\bfarm\w*\b', r'\bagricultur\w*\b', r'\bgrain\b', r'\bwheat\b',
    r'\bsoybeans?\b', r'\bfertilizer\b',
    r'\bpharmaceutical\b', r'\bdrug\s+(supply|shortage)\b',
    r'\bfood\s+(supply|chain|shortage)\b', r'\bfood\s+processing\b',
    r'\bretail\b', r'\becommerce\b', r'\bonline\s+shopping\b',
    r'\bfurniture\b', r'\btextile\b', r'\bapparel\b', r'\bfashion\b',
]


def count_hits(text, patterns):
    t = text.lower()
    return sum(1 for p in patterns if re.search(p, t, re.IGNORECASE))


def classify_v2(title: str, content: str) -> tuple[int, str]:
    """
    Returns (label, reason_string)
    """
    txt_full   = f"{title} {title} {title} {content}"  # title weight x3
    title_l    = title.lower()
    content_l  = content.lower()
    full_l     = txt_full.lower()

    # --- Hard neutral check (technology/admin articles) ----------------------
    neutral_hits = count_hits(full_l, NEUTRAL_HARD)
    if neutral_hits >= 2:
        return 0, f"NEUTRAL_HARD ({neutral_hits} hits)"

    # --- Check aerospace scope -----------------------------------------------
    aerospace_hits    = count_hits(full_l, AEROSPACE_RELATED)
    non_aero_hits     = count_hits(full_l, NON_AEROSPACE_HARD)
    # QT-03 ap dung khi: ro rang la nganh khac, khong co tu khoa aerospace
    apply_qt03 = (non_aero_hits >= 1 and aerospace_hits == 0)

    # --- Score HIGH RISK (su kien DA xay ra) ---------------------------------
    h_title  = count_hits(title_l, ALL_HIGH_PATTERNS)
    h_full   = count_hits(full_l,  ALL_HIGH_PATTERNS)
    h_score  = h_title * 3 + h_full

    # Sub-scores theo tung category de debug
    dbg = {
        "strike_active":    count_hits(full_l, STRIKE_ACTIVE),
        "red_sea_active":   count_hits(full_l, RED_SEA_ACTIVE),
        "port_closed":      count_hits(full_l, PORT_CLOSED),
        "plant_closed":     count_hits(full_l, PLANT_CLOSED),
        "bankruptcy":       count_hits(full_l, BANKRUPTCY_ACTIVE),
        "disaster_active":  count_hits(full_l, DISASTER_ACTIVE),
        "sanction_active":  count_hits(full_l, SANCTION_ACTIVE),
        "lockdown":         count_hits(full_l, LOCKDOWN_ACTIVE),
        "shortage_conf":    count_hits(full_l, SHORTAGE_CONFIRMED),
        "cargo_seized":     count_hits(full_l, CARGO_SEIZED),
        "severe_delay":     count_hits(full_l, SEVERE_DELAY),
    }
    top_cat = max(dbg, key=dbg.get)

    # --- Score MEDIUM RISK (dau hieu dang hinh thanh) -----------------------
    m_title  = count_hits(title_l, MEDIUM_PATTERNS)
    m_full   = count_hits(full_l,  MEDIUM_PATTERNS)
    m_score  = m_title * 2 + m_full

    # --- QT-04: Bai dung ngon ngu manh nhung su kien chua xay ra ------------
    # Phat hien "su kien tuong lai" pattern -- neu co thi cap che High score
    FUTURE_HEDGE = [
        r'\b(could|may|might|would|should|expect\w*|forecast\w*|predict\w*|anticipat\w*|project\w*)\b',
        r'\b(if|unless|in\s+case|should\s+\w+\s+happen)\b',
        r'\b(risk\s+of|threat\s+of|fear\s+of|concern\s+about)\b',
        r'\b(warn\w*|alert\w*|caution\w*)\s+\w{0,20}\s+(could|may|might)\b',
        r'\blooming\b', r'\bpotential\b', r'\bpossible\b', r'\bpending\b',
    ]
    future_hedge_hits = count_hits(title_l + " " + content_l[:500], FUTURE_HEDGE)

    # --- Decision logic ------------------------------------------------------
    raw_label = 0
    reason    = "NO_RISK default"

    if h_score >= 5 or h_title >= 2:
        raw_label = 2
        reason = f"HIGH: h_score={h_score} h_title={h_title} top_cat={top_cat}"
    elif h_score >= 3 and future_hedge_hits <= 2:
        raw_label = 2
        reason = f"HIGH: h_score={h_score} hedge={future_hedge_hits} top_cat={top_cat}"
    elif h_score >= 2 and m_score >= 2 and future_hedge_hits <= 1:
        raw_label = 2
        reason = f"HIGH: h={h_score} m={m_score} hedge={future_hedge_hits}"
    elif h_score >= 2 or (h_score == 1 and m_score >= 4):
        raw_label = 1
        reason = f"MEDIUM: h={h_score} m={m_score} (borderline)"
    elif m_score >= 4:
        raw_label = 1
        reason = f"MEDIUM: m_score={m_score}"
    elif m_score >= 2 and neutral_hits == 0:
        raw_label = 1
        reason = f"MEDIUM: m_score={m_score}"
    elif m_score >= 1 and aerospace_hits >= 1:
        raw_label = 1
        reason = f"MEDIUM: m={m_score} aerospace_hits={aerospace_hits}"
    else:
        raw_label = 0
        reason = f"NO_RISK: h={h_score} m={m_score} neutral={neutral_hits}"

    # --- QT-03: Giam 1 bac neu khong lien quan aerospace ---------------------
    if apply_qt03 and raw_label > 0:
        old = raw_label
        raw_label = max(0, raw_label - 1)
        reason += f" | QT-03 applied ({old}->{raw_label}, non_aero={non_aero_hits})"

    return raw_label, reason


# ── Apply to all 150 articles ─────────────────────────────────────────────────
print("Classifying 150 articles with v2 guide...")
results = []
for row in all_150:
    art = row["art"]
    lbl, reason = classify_v2(
        art.get("title", "") or "",
        art.get("content","") or "",
    )
    results.append({"idx": row["idx"], "block": row["block"],
                    "label": lbl, "reason": reason,
                    "title": art.get("title","")[:90]})
    print(f"  [{row['idx']:03d}] L={lbl} | {art.get('title','')[:75]}")
    if "QT-03" in reason or lbl == 2:
        print(f"         -> {reason}")

dist = {0: sum(1 for r in results if r["label"]==0),
        1: sum(1 for r in results if r["label"]==1),
        2: sum(1 for r in results if r["label"]==2)}
print(f"\nDistribution v2: {dist}  (total={len(results)})")

# ══════════════════════════════════════════════════════════════════════════════
# UPDATE EXCEL: xoa nhan cu, dien lai cot Claude, xoa cot H/I/J
# ══════════════════════════════════════════════════════════════════════════════
def fill_cell(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

LABEL_BG = {0: "E2EFDA", 1: "FFF2CC", 2: "FCE4D6"}
LABEL_FG = {0: "375623", 1: "7F6000", 2: "843C0C"}
EMPTY_BG = "FFF2CC"   # mau nen cho o trong cua rater khac

print(f"\nOpening: {XL}")
wb = openpyxl.load_workbook(XL)
ws = wb["📋 Gán Nhãn"]

# Build mapping: idx -> label (skip separator row)
label_map = {r["idx"]: r["label"] for r in results}

# Duyet tung row data (row 3 den 153, bo separator)
for excel_row in range(3, 154):
    # Kiem tra merged cell - skip neu la MergedCell
    a_cell = ws[f"A{excel_row}"]
    if a_cell.__class__.__name__ == "MergedCell":
        continue
    stt_val = a_cell.value
    if stt_val is None or not str(stt_val).strip().isdigit():
        # Separator row hoac row trong -> skip, khong can xoa
        continue

    idx = int(str(stt_val).strip())
    lbl = label_map.get(idx, 0)

    # Cot G: dien nhan Claude v2
    c_g = ws[f"G{excel_row}"]
    c_g.value     = lbl
    c_g.font      = Font(bold=True, color=LABEL_FG[lbl], size=11, name="Calibri")
    c_g.fill      = fill_cell(LABEL_BG[lbl])
    c_g.alignment = Alignment(horizontal="center", vertical="center")
    c_g.border    = thin_border()

    # Cot H, I, J: xoa nhan cu, giu mau nen vang nhat
    for col in ["H", "I", "J"]:
        c = ws[f"{col}{excel_row}"]
        c.value     = None
        c.font      = Font(size=10, name="Calibri")
        c.fill      = fill_cell(EMPTY_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

wb.save(XL)
print(f"Saved: {XL}")
print(f"Label v2 distribution: 0={dist[0]}  1={dist[1]}  2={dist[2]}")
print(f"Label 2 rate: {dist[2]/150*100:.1f}%")
